"""
Output Generator - Generate Excel/CSV files for import into accounting systems
Creates formatted output files for Cash Receipts, Cash Disbursements, Journal Vouchers
"""

import os
import sys
from typing import Dict, List, Optional
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import OUTPUT_DIR, OUTPUT_FILES

class OutputGenerator:
    """Generate formatted output files for accounting system import"""
    
    def __init__(self, output_dir: str = None, target_system: str = 'MIP'):
        self.output_dir = output_dir or OUTPUT_DIR
        self.target_system = target_system
        os.makedirs(self.output_dir, exist_ok=True)
        self.generated_files = []
        
    def generate_all(self, entries: Dict, timestamp: str = None) -> Dict:
        """Generate all output files"""
        if timestamp is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        files = {}
        
        if entries.get('CR'):
            files['CR'] = self.generate_cash_receipts(entries['CR'], timestamp)
        if entries.get('CD'):
            files['CD'] = self.generate_cash_disbursements(entries['CD'], timestamp)
        if entries.get('JV'):
            files['JV'] = self.generate_journal_vouchers(entries['JV'], timestamp)
        
        return files
    
    def generate_cash_receipts(self, entries: List[Dict], timestamp: str) -> str:
        """Generate Cash Receipts import file"""
        import pandas as pd
        
        filename = f"Cash_Receipts_Import_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        rows = []
        for entry in entries:
            for line in entry.get('lines', []):
                rows.append({
                    'Session ID': entry.get('session_id', ''),
                    'Doc Number': entry.get('doc_number', ''),
                    'Doc Date': entry.get('doc_date', ''),
                    'Payer Name': entry.get('payer_name', ''),
                    'Receipt Type': entry.get('receipt_type', ''),
                    'GL Code': line.get('gl_code', ''),
                    'Fund Code': line.get('fund_code', ''),
                    'Debit': line.get('debit', 0) if line.get('debit', 0) > 0 else '',
                    'Credit': line.get('credit', 0) if line.get('credit', 0) > 0 else '',
                    'Description': entry.get('description', ''),
                    'Reference': entry.get('doc_number', ''),
                    'Needs Review': 'Yes' if entry.get('needs_review') else ''
                })
        
        df = pd.DataFrame(rows)
        self._save_with_formatting(df, filepath, 'Cash Receipts')
        self.generated_files.append(filepath)
        return filepath
    
    def generate_cash_disbursements(self, entries: List[Dict], timestamp: str) -> str:
        """Generate Cash Disbursements import file"""
        import pandas as pd
        
        filename = f"Cash_Disbursements_Import_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        rows = []
        for entry in entries:
            for line in entry.get('lines', []):
                rows.append({
                    'Session ID': entry.get('session_id', ''),
                    'Doc Number': entry.get('doc_number', ''),
                    'Doc Date': entry.get('doc_date', ''),
                    'Vendor Name': entry.get('vendor_name', ''),
                    'Payment Method': entry.get('payment_method', ''),
                    'Check Number': entry.get('check_number', ''),
                    'Disbursement Type': entry.get('disbursement_type', ''),
                    'GL Code': line.get('gl_code', ''),
                    'Fund Code': line.get('fund_code', ''),
                    'Debit': line.get('debit', 0) if line.get('debit', 0) > 0 else '',
                    'Credit': line.get('credit', 0) if line.get('credit', 0) > 0 else '',
                    'Description': entry.get('description', ''),
                    'Reference': entry.get('reference', ''),
                    'Needs Review': 'Yes' if entry.get('needs_review') else ''
                })
        
        df = pd.DataFrame(rows)
        self._save_with_formatting(df, filepath, 'Cash Disbursements')
        self.generated_files.append(filepath)
        return filepath
    
    def generate_journal_vouchers(self, entries: List[Dict], timestamp: str) -> str:
        """Generate Journal Vouchers import file"""
        import pandas as pd
        
        filename = f"Journal_Vouchers_Import_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        rows = []
        for entry in entries:
            for line in entry.get('lines', []):
                rows.append({
                    'Session ID': entry.get('session_id', ''),
                    'Doc Number': entry.get('doc_number', ''),
                    'Doc Date': entry.get('doc_date', ''),
                    'JV Type': entry.get('jv_type', ''),
                    'GL Code': line.get('gl_code', ''),
                    'Fund Code': line.get('fund_code', ''),
                    'Debit': line.get('debit', 0) if line.get('debit', 0) > 0 else '',
                    'Credit': line.get('credit', 0) if line.get('credit', 0) > 0 else '',
                    'Description': entry.get('description', ''),
                    'Reference': entry.get('reference', ''),
                    'Needs Review': 'Yes' if entry.get('needs_review') else ''
                })
        
        df = pd.DataFrame(rows)
        self._save_with_formatting(df, filepath, 'Journal Vouchers')
        self.generated_files.append(filepath)
        return filepath
    
    def generate_unidentified(self, transactions: List[Dict], timestamp: str) -> str:
        """Generate Unidentified transactions file"""
        import pandas as pd
        
        filename = f"Unidentified_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        rows = []
        for txn in transactions:
            rows.append({
                'Date': txn.get('date', ''),
                'Description': txn.get('description', ''),
                'Amount': txn.get('amount', 0),
                'Suggested Module': '',
                'GL Code': '',
                'Fund Code': '',
                'Vendor/Customer': '',
                'Notes': 'Requires manual classification'
            })
        
        df = pd.DataFrame(rows)
        self._save_with_formatting(df, filepath, 'Unidentified', highlight_all=True)
        self.generated_files.append(filepath)
        return filepath
    
    def generate_summary_report(self, entries: Dict, classification_summary: Dict, 
                                routing_summary: Dict, timestamp: str) -> str:
        """Generate summary report"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        filename = f"Processing_Summary_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        
        ws['A1'] = "Bank Transaction Posting Tool - Processing Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        ws['A4'] = "Classification Summary"
        ws['A4'].font = Font(bold=True)
        row = 5
        ws[f'A{row}'] = "Total Transactions"
        ws[f'B{row}'] = classification_summary.get('total', 0)
        
        row += 1
        ws[f'A{row}'] = "By Module:"
        for module, count in classification_summary.get('by_module', {}).items():
            row += 1
            ws[f'A{row}'] = f"  {module}"
            ws[f'B{row}'] = count
        
        row += 2
        ws[f'A{row}'] = "Amount Summary"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "Total Credits"
        ws[f'B{row}'] = f"${classification_summary.get('total_credits', 0):,.2f}"
        row += 1
        ws[f'A{row}'] = "Total Debits"
        ws[f'B{row}'] = f"${classification_summary.get('total_debits', 0):,.2f}"
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        wb.save(filepath)
        self.generated_files.append(filepath)
        return filepath
    
    def _save_with_formatting(self, df, filepath: str, sheet_name: str, highlight_all: bool = False):
        """Save DataFrame with professional formatting"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 1:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                elif highlight_all or (c_idx == len(row) and value == 'Yes'):
                    if value == 'Yes' or highlight_all:
                        for col in range(1, len(row) + 1):
                            ws.cell(row=r_idx, column=col).fill = PatternFill(
                                start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'
        wb.save(filepath)
    
    def get_generated_files(self) -> List[str]:
        return self.generated_files


if __name__ == "__main__":
    generator = OutputGenerator()
    
    test_entries = {
        'CR': [{
            'session_id': 'GP_CR_2024', 'doc_number': 'GP_1201_001', 'doc_date': '12/01/2024',
            'payer_name': 'HUD', 'receipt_type': 'Grant Receipt', 'description': 'HUD Grant',
            'lines': [
                {'gl_code': '1070', 'fund_code': '2700', 'debit': 50000, 'credit': 0},
                {'gl_code': '4100', 'fund_code': '2700', 'debit': 0, 'credit': 50000}
            ], 'needs_review': False
        }],
        'CD': [{
            'session_id': 'GP_CD_2024', 'doc_number': 'GP_1201_001', 'doc_date': '12/01/2024',
            'vendor_name': 'ADP', 'payment_method': 'ACH', 'disbursement_type': 'Payroll',
            'description': 'Payroll', 'reference': 'Dec',
            'lines': [
                {'gl_code': '7200', 'fund_code': '2700', 'debit': 15000, 'credit': 0},
                {'gl_code': '1070', 'fund_code': '2700', 'debit': 0, 'credit': 15000}
            ], 'needs_review': False
        }],
        'JV': [{
            'session_id': 'GP_JV_2024', 'doc_number': 'GP_1201_001', 'doc_date': '12/01/2024',
            'jv_type': 'Bank Fee', 'description': 'Service Charge', 'reference': 'Dec',
            'lines': [
                {'gl_code': '7500', 'fund_code': '1000', 'debit': 25, 'credit': 0},
                {'gl_code': '1070', 'fund_code': '1000', 'debit': 0, 'credit': 25}
            ], 'needs_review': False
        }]
    }
    
    files = generator.generate_all(test_entries)
    print("Generated Files:", files)
