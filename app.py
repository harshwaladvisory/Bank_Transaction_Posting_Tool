"""
Bank Transaction Posting Tool - Enhanced Flask Web Interface
Features: Autocomplete, Confidence Colors, Bulk Actions, Audit Trail, Vendor Matching
         Customer/Vendor Management with Auto-fill GL
         Complete Chart of Accounts from Client
         MongoDB Integration with REST API Endpoints
"""

import os
import sys
import json
import io
import base64
from datetime import datetime
from functools import wraps
from bson import ObjectId
from flask import Flask, render_template_string, request, redirect, url_for, flash, jsonify, send_file, Response, make_response
from flask import session  # Explicit session import
from werkzeug.utils import secure_filename

# MongoDB imports
try:
    from pymongo import MongoClient
    from pymongo.errors import ConnectionFailure, ServerSelectionTimeoutError
    MONGODB_AVAILABLE = True
except ImportError:
    MONGODB_AVAILABLE = False
    print("WARNING: pymongo not installed. Run: pip install pymongo")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import DATA_DIR
from parsers import UniversalParser
from classifiers import ClassificationEngine
from processors import ModuleRouter, EntryBuilder, OutputGenerator

app = Flask(__name__)

# SECURITY FIX: Use environment variable for secret key
import secrets as secrets_module
app.secret_key = os.environ.get('SECRET_KEY', secrets_module.token_hex(32))

# Warn if using default secret key
if 'SECRET_KEY' not in os.environ:
    print("[WARNING] SECRET_KEY environment variable not set!")
    print("   Using auto-generated key. Sessions will not persist across restarts.")
    print("   For production, set: export SECRET_KEY='your-secret-key-here'")

app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Disable caching for all responses to ensure fresh data after uploads
@app.after_request
def add_no_cache_headers(response):
    """Add no-cache headers to all responses to prevent stale data"""
    if 'text/html' in response.content_type or response.content_type == 'application/json':
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# MongoDB Configuration
MONGODB_URI = os.environ.get('MONGODB_URI', 'mongodb://localhost:27017/')
MONGODB_DATABASE = os.environ.get('MONGODB_DATABASE', 'bank_posting_tool')

# Note: Files are now processed in-memory, no local upload or output directory needed
# REMOVED: Global session_data (replaced with per-user sessions in MongoDB)

# ============ MONGODB CONNECTION ============

# Global MongoDB client (initialized lazily)
_mongo_client = None
_mongo_db = None
_mongo_connection_failed = False

def get_db():
    """Get MongoDB database connection (lazy initialization with caching and retry)"""
    global _mongo_client, _mongo_db, _mongo_connection_failed

    if not MONGODB_AVAILABLE:
        return None

    # If connection previously failed, retry every 5 minutes
    if _mongo_connection_failed:
        if not hasattr(get_db, 'last_retry_time'):
            get_db.last_retry_time = datetime.now()

        # Check if 5 minutes have passed since last retry
        time_since_last_retry = (datetime.now() - get_db.last_retry_time).total_seconds()
        if time_since_last_retry > 300:  # 5 minutes
            print("Retrying MongoDB connection (5 minutes since last attempt)...")
            _mongo_connection_failed = False  # Allow retry
            get_db.last_retry_time = datetime.now()
        else:
            return None

    # Return cached connection if available
    if _mongo_db is not None:
        try:
            # Test connection is still alive
            _mongo_client.admin.command('ping')
            return _mongo_db
        except:
            # Connection lost, retry
            print("MongoDB connection lost, reconnecting...")
            _mongo_db = None
            _mongo_client = None

    # Try to establish connection (only once at startup or first request)
    try:
        _mongo_client = MongoClient(
            MONGODB_URI,
            serverSelectionTimeoutMS=2000,  # Reduced from 5000ms to 2000ms
            connectTimeoutMS=2000,
            socketTimeoutMS=2000
        )
        # Test connection with quick timeout
        _mongo_client.admin.command('ping')
        _mongo_db = _mongo_client[MONGODB_DATABASE]
        print(f"[OK] MongoDB connected: {MONGODB_URI}{MONGODB_DATABASE}")
        return _mongo_db
    except (ConnectionFailure, ServerSelectionTimeoutError) as e:
        print(f"[ERROR] MongoDB connection failed: {e}")
        print(f"  -> Application will run with static data only")
        _mongo_connection_failed = True
        return None
    except Exception as e:
        print(f"[ERROR] MongoDB error: {e}")
        _mongo_connection_failed = True
        return None

def init_mongodb():
    """Initialize MongoDB collections and indexes (lazy, non-blocking)"""
    db = get_db()
    if db is None:
        print("  -> Application running without MongoDB (using static data)")
        return False

    try:
        # Create collections if they don't exist
        collections = ['transactions', 'gl_codes', 'fund_codes', 'vendors', 'customers', 'batches', 'audit_logs', 'output_files']
        existing = db.list_collection_names()
        for coll in collections:
            if coll not in existing:
                db.create_collection(coll)

        # Create indexes (with background=True to avoid blocking)
        db.transactions.create_index([('batch_id', 1)], background=True)
        db.transactions.create_index([('date', -1)], background=True)
        db.transactions.create_index([('module', 1)], background=True)
        db.transactions.create_index([('status', 1)], background=True)
        db.gl_codes.create_index([('code', 1)], unique=True, background=True)
        db.fund_codes.create_index([('code', 1)], unique=True, background=True)
        db.vendors.create_index([('name', 1)], background=True)
        db.customers.create_index([('name', 1)], background=True)
        db.batches.create_index([('created_at', -1)], background=True)
        db.audit_logs.create_index([('timestamp', -1)], background=True)
        db.output_files.create_index([('created_at', -1)], background=True)
        db.output_files.create_index([('batch_id', 1)], background=True)

        print(f"  -> MongoDB collections and indexes initialized")
        return True
    except Exception as e:
        print(f"  -> MongoDB initialization warning: {e}")
        # Don't return False - connection works, just initialization had issues
        return True

def serialize_doc(doc):
    """Convert MongoDB document to JSON-serializable dict"""
    if doc is None:
        return None
    if isinstance(doc, list):
        return [serialize_doc(d) for d in doc]
    if isinstance(doc, dict):
        result = {}
        for k, v in doc.items():
            if isinstance(v, ObjectId):
                result[k] = str(v)
            elif isinstance(v, datetime):
                result[k] = v.isoformat()
            elif isinstance(v, dict):
                result[k] = serialize_doc(v)
            elif isinstance(v, list):
                result[k] = serialize_doc(v)
            else:
                result[k] = v
        return result
    return doc

# MongoDB will be initialized lazily on first use (non-blocking startup)
# This prevents the app from hanging if MongoDB is unavailable
mongodb_ready = False

# ============ COMPLETE CHART OF ACCOUNTS FROM CLIENT ============

GL_CODES = [
    # Bank Accounts
    {'code': '10000', 'name': 'Cash', 'type': 'Bank'},
    {'code': '10104', 'name': 'PNC - Operating Checking', 'type': 'Bank'},
    {'code': '10105', 'name': 'PNC - Business Checking - 8729', 'type': 'Bank'},
    {'code': '10109', 'name': 'PNC - Business Checking - 8788', 'type': 'Bank'},
    {'code': '10110', 'name': 'PNC - Construction Acct', 'type': 'Bank'},
    {'code': '10111', 'name': 'PNC - Business Checking - 8833', 'type': 'Bank'},
    {'code': '10112', 'name': 'TRUIST-Community Checking 0719', 'type': 'Bank'},
    {'code': '10113', 'name': 'TRUIST - Larger MM 6505', 'type': 'Bank'},
    {'code': '10114', 'name': 'TRUIST - Pow-Wow MM 6513', 'type': 'Bank'},
    {'code': '10115', 'name': 'TRUIST - BGCA 6483', 'type': 'Bank'},
    {'code': '10116', 'name': 'TRUIST - Pow-Wow 6491', 'type': 'Bank'},
    {'code': '10120', 'name': 'Pinnacle Bank - Checking', 'type': 'Bank'},
    {'code': '10202', 'name': 'Edward Jones Investment', 'type': 'Bank'},
    {'code': '10204', 'name': 'PNC - MM', 'type': 'Bank'},
    {'code': '10490', 'name': 'Bill.com Money Out Clearing', 'type': 'Bank'},
    {'code': '10491', 'name': 'BillCom Money In Clearing', 'type': 'Bank'},
    
    # Receivables
    {'code': '11000', 'name': 'Accounts Receivable', 'type': 'Asset'},
    {'code': '11001', 'name': 'Account Receivable Other', 'type': 'Asset'},
    {'code': '12000', 'name': 'Undeposited Funds', 'type': 'Asset'},
    {'code': '12100', 'name': 'Inventory Asset', 'type': 'Asset'},
    {'code': '12500', 'name': 'Due From', 'type': 'Asset'},
    {'code': '13000', 'name': 'Prepaid Expenses', 'type': 'Asset'},
    
    # Payables & Liabilities
    {'code': '20000', 'name': 'Accounts Payable', 'type': 'Liability'},
    {'code': '21000', 'name': 'Deferred Revenue', 'type': 'Liability'},
    {'code': '22500', 'name': 'Due To', 'type': 'Liability'},
    {'code': '23000', 'name': 'Accrued Vacation', 'type': 'Liability'},
    {'code': '24000', 'name': 'Payroll Liabilities', 'type': 'Liability'},
    {'code': '24010', 'name': '403(b) Deferrals', 'type': 'Liability'},
    {'code': '24020', 'name': 'AFLAC Supplemental Insurance', 'type': 'Liability'},
    {'code': '24030', 'name': 'Federal Withholding', 'type': 'Liability'},
    {'code': '24040', 'name': 'FICA/Medicare Withholding', 'type': 'Liability'},
    {'code': '24050', 'name': 'Garnishments/Orders', 'type': 'Liability'},
    {'code': '24060', 'name': 'Health Insurance', 'type': 'Liability'},
    {'code': '24070', 'name': 'NC Unemployment', 'type': 'Liability'},
    {'code': '24080', 'name': 'NC Withholding', 'type': 'Liability'},
    {'code': '24090', 'name': 'Other Withholdings', 'type': 'Liability'},
    {'code': '24100', 'name': 'Simple IRA', 'type': 'Liability'},
    {'code': '25000', 'name': 'Other Current Liabilities', 'type': 'Liability'},
    
    # Revenue - Federal/State
    {'code': '3001', 'name': 'Revenue - Federal', 'type': 'Revenue'},
    {'code': '3002', 'name': 'Revenue - State', 'type': 'Revenue'},
    {'code': '4000', 'name': 'Revenue - Local', 'type': 'Revenue'},
    {'code': '4001', 'name': 'Revenue - Youth Empowerment', 'type': 'Revenue'},
    {'code': '4003', 'name': 'Revenue - Various', 'type': 'Revenue'},
    {'code': '4010', 'name': 'Revenue - Drinks', 'type': 'Revenue'},
    {'code': '4070', 'name': 'Revenue - Appropriations', 'type': 'Revenue'},
    {'code': '4080', 'name': 'Revenue - Contributions/Donations', 'type': 'Revenue'},
    {'code': '4090', 'name': 'Revenue - Day Camp', 'type': 'Revenue'},
    {'code': '4100', 'name': 'Revenue - Fundraising', 'type': 'Revenue'},
    {'code': '4110', 'name': 'Revenue - Mutual Help/Emergency', 'type': 'Revenue'},
    {'code': '4120', 'name': 'Revenue - Tribal Enrollment', 'type': 'Revenue'},
    {'code': '4125', 'name': 'Revenue - Tribal Meeting', 'type': 'Revenue'},
    {'code': '4130', 'name': 'Revenue - CPTA', 'type': 'Revenue'},
    {'code': '4140', 'name': 'Revenue - Federal Recognition', 'type': 'Revenue'},
    {'code': '4150', 'name': 'Revenue - Princess Committee', 'type': 'Revenue'},
    {'code': '4160', 'name': 'Revenue - Cultural Exchange', 'type': 'Revenue'},
    {'code': '4180', 'name': 'Revenue - Youth Leadership', 'type': 'Revenue'},
    {'code': '4215', 'name': 'Revenue - Blaylock Scholarship', 'type': 'Revenue'},
    {'code': '4220', 'name': 'Revenue - WR Scholarship', 'type': 'Revenue'},
    {'code': '4225', 'name': 'Revenue - Scholarship', 'type': 'Revenue'},
    {'code': '4240', 'name': 'Revenue - NAOP Grant', 'type': 'Revenue'},
    {'code': '4250', 'name': 'Revenue - URP Program', 'type': 'Revenue'},
    {'code': '4260', 'name': 'Revenue - Membership', 'type': 'Revenue'},
    {'code': '4300', 'name': 'Revenue - EVENTS', 'type': 'Revenue'},
    {'code': '4301', 'name': 'Revenue - Advertising', 'type': 'Revenue'},
    {'code': '4302', 'name': 'Revenue - Gate', 'type': 'Revenue'},
    {'code': '4303', 'name': 'Revenue - Food', 'type': 'Revenue'},
    {'code': '4304', 'name': 'Revenue - Information', 'type': 'Revenue'},
    {'code': '4305', 'name': 'Revenue - Drinks', 'type': 'Revenue'},
    {'code': '4306', 'name': 'Revenue - Vendor', 'type': 'Revenue'},
    {'code': '4315', 'name': 'Revenue - Rides', 'type': 'Revenue'},
    {'code': '4325', 'name': 'Revenue - Insurance Revenue', 'type': 'Revenue'},
    {'code': '4330', 'name': 'Revenue - Miscellaneous', 'type': 'Revenue'},
    {'code': '4380', 'name': 'Revenue - Tomahawk Meadows', 'type': 'Revenue'},
    {'code': '4390', 'name': 'Revenue - Red Earth Village', 'type': 'Revenue'},
    {'code': '4410', 'name': 'Revenue - Daycare Fees', 'type': 'Revenue'},
    {'code': '4479', 'name': 'Indirect Revenues', 'type': 'Revenue'},
    {'code': '4720', 'name': 'Revenue - TAPS', 'type': 'Revenue'},
    
    # Rent Revenue
    {'code': '4020', 'name': 'Revenue - Rent (School)', 'type': 'Revenue'},
    {'code': '4035', 'name': 'Revenue - Rent (Other)', 'type': 'Revenue'},
    {'code': '4040', 'name': 'Revenue - Rent (Library)', 'type': 'Revenue'},
    {'code': '4050', 'name': 'Revenue - Rent (Tribal Center)', 'type': 'Revenue'},
    {'code': '4060', 'name': 'Revenue - Rent (Mobile Home)', 'type': 'Revenue'},
    {'code': '4062', 'name': 'Revenue - Rent (House)', 'type': 'Revenue'},
    {'code': '4065', 'name': 'Revenue - Rent (Duplex)', 'type': 'Revenue'},
    
    # Other Income
    {'code': '9010', 'name': 'Interest Income', 'type': 'Other Income'},
    {'code': '9020', 'name': 'Dividend Income', 'type': 'Other Income'},
    
    # Expenses - General Operating
    {'code': '5010', 'name': 'Advertising', 'type': 'Expense'},
    {'code': '5015', 'name': 'Promotional Products/Outreach', 'type': 'Expense'},
    {'code': '5020', 'name': 'Assistance - Mortgage', 'type': 'Expense'},
    {'code': '5025', 'name': 'Assistance - Rental', 'type': 'Expense'},
    {'code': '5030', 'name': 'Youth Leadership', 'type': 'Expense'},
    {'code': '5050', 'name': 'Audit Fees', 'type': 'Expense'},
    {'code': '5060', 'name': 'Bank Charges', 'type': 'Expense'},
    {'code': '5064', 'name': 'Boys and Girls Club', 'type': 'Expense'},
    {'code': '5065', 'name': 'Capital Outlay Expense', 'type': 'Expense'},
    {'code': '5070', 'name': 'Childcare Subsidy', 'type': 'Expense'},
    {'code': '5085', 'name': 'Community/Recreation Center', 'type': 'Expense'},
    {'code': '5086', 'name': 'National Night Out', 'type': 'Expense'},
    {'code': '5100', 'name': 'Consultant', 'type': 'Expense'},
    {'code': '5105', 'name': 'Speaker Fees', 'type': 'Expense'},
    {'code': '5110', 'name': 'Contracted Services', 'type': 'Expense'},
    {'code': '5115', 'name': 'Contributions/Donations', 'type': 'Expense'},
    {'code': '5120', 'name': 'Cultural Enrichment', 'type': 'Expense'},
    {'code': '5128', 'name': 'Competition - Art', 'type': 'Expense'},
    {'code': '5130', 'name': 'Competition - Dance/Drum', 'type': 'Expense'},
    {'code': '5137', 'name': 'Debt Service - Principal', 'type': 'Expense'},
    {'code': '5138', 'name': 'Debt Service - Interest', 'type': 'Expense'},
    {'code': '5150', 'name': 'Drinks & Ice', 'type': 'Expense'},
    {'code': '5152', 'name': 'Software - Accounting and other', 'type': 'Expense'},
    {'code': '5170', 'name': 'Emergency Assistance', 'type': 'Expense'},
    {'code': '5180', 'name': 'Equipment (Non Capitalized)', 'type': 'Expense'},
    {'code': '5181', 'name': 'HVAC Upgrade', 'type': 'Expense'},
    {'code': '5185', 'name': 'Sports Equipment', 'type': 'Expense'},
    {'code': '5190', 'name': 'Federal Recognition', 'type': 'Expense'},
    {'code': '5200', 'name': 'Fundraising Expense', 'type': 'Expense'},
    {'code': '5220', 'name': 'CPTA', 'type': 'Expense'},
    {'code': '5230', 'name': 'Homebuyers Assistance', 'type': 'Expense'},
    {'code': '5240', 'name': 'Host Drum', 'type': 'Expense'},
    {'code': '5260', 'name': 'Improvement', 'type': 'Expense'},
    {'code': '5270', 'name': 'Indirect Charged to Grants', 'type': 'Expense'},
    {'code': '5280', 'name': 'Insurance - General', 'type': 'Expense'},
    {'code': '5320', 'name': 'Insurance - Vehicle', 'type': 'Expense'},
    {'code': '5325', 'name': 'Internet', 'type': 'Expense'},
    {'code': '5326', 'name': 'IT/Services', 'type': 'Expense'},
    {'code': '5330', 'name': 'Land Development', 'type': 'Expense'},
    {'code': '5392', 'name': 'Appliance Replacement', 'type': 'Expense'},
    {'code': '5480', 'name': 'Head Staff', 'type': 'Expense'},
    {'code': '5490', 'name': 'Medical Examinations', 'type': 'Expense'},
    {'code': '5495', 'name': 'Drug Screening', 'type': 'Expense'},
    {'code': '5500', 'name': 'Membership Fees', 'type': 'Expense'},
    {'code': '5505', 'name': 'Registration Fees', 'type': 'Expense'},
    {'code': '5510', 'name': 'Other', 'type': 'Expense'},
    {'code': '5540', 'name': 'Penalties', 'type': 'Expense'},
    {'code': '5570', 'name': 'Postage', 'type': 'Expense'},
    {'code': '5580', 'name': 'Princess Committee', 'type': 'Expense'},
    {'code': '5590', 'name': 'Printing/Publications', 'type': 'Expense'},
    {'code': '5600', 'name': 'Professional Srvs - Accounting', 'type': 'Expense'},
    {'code': '5610', 'name': 'Professional Srvs - Attorney', 'type': 'Expense'},
    {'code': '5615', 'name': 'Back Ground Checks', 'type': 'Expense'},
    {'code': '5630', 'name': 'Property Taxes', 'type': 'Expense'},
    {'code': '5635', 'name': 'Property Taxes - Solid Waste', 'type': 'Expense'},
    {'code': '5640', 'name': 'Public Relations', 'type': 'Expense'},
    {'code': '5645', 'name': '4th of July', 'type': 'Expense'},
    {'code': '5650', 'name': 'Daycare Renovation', 'type': 'Expense'},
    {'code': '5660', 'name': 'Rehabilitation', 'type': 'Expense'},
    {'code': '5662', 'name': 'Rehabilitation - Emergency', 'type': 'Expense'},
    {'code': '5665', 'name': 'Reimbursement - General', 'type': 'Expense'},
    {'code': '5695', 'name': 'Conference Room/Space Rental', 'type': 'Expense'},
    {'code': '5700', 'name': 'Rent Expense', 'type': 'Expense'},
    {'code': '5710', 'name': 'Rent - Other', 'type': 'Expense'},
    {'code': '5730', 'name': 'Sales Tax', 'type': 'Expense'},
    {'code': '5740', 'name': 'Scholarship', 'type': 'Expense'},
    {'code': '5741', 'name': 'Scholarship Farrington', 'type': 'Expense'},
    {'code': '5742', 'name': 'College Housing Assistance', 'type': 'Expense'},
    {'code': '5750', 'name': 'Security', 'type': 'Expense'},
    {'code': '5815', 'name': 'Incentives', 'type': 'Expense'},
    {'code': '5816', 'name': 'Participant Cost', 'type': 'Expense'},
    {'code': '5830', 'name': 'Telephone', 'type': 'Expense'},
    {'code': '5840', 'name': 'Training - Classroom', 'type': 'Expense'},
    {'code': '5844', 'name': 'Staff Uniforms', 'type': 'Expense'},
    {'code': '5845', 'name': 'Staff Development', 'type': 'Expense'},
    {'code': '5846', 'name': 'Traditional Clothing', 'type': 'Expense'},
    {'code': '5847', 'name': 'Non-Traditional Clothing', 'type': 'Expense'},
    {'code': '5850', 'name': 'Training - Other', 'type': 'Expense'},
    {'code': '5855', 'name': 'Gas/Fuel', 'type': 'Expense'},
    {'code': '5860', 'name': 'Travel - Local', 'type': 'Expense'},
    {'code': '5865', 'name': 'Gas Vouchers', 'type': 'Expense'},
    {'code': '5870', 'name': 'Travel - Out of Area', 'type': 'Expense'},
    {'code': '5876', 'name': 'Field Trips', 'type': 'Expense'},
    {'code': '5880', 'name': 'Tribal Enrollment', 'type': 'Expense'},
    {'code': '5890', 'name': 'Tribal Princess', 'type': 'Expense'},
    {'code': '5950', 'name': 'Waste Disposal', 'type': 'Expense'},
    {'code': '5960', 'name': 'Youth Day Camp', 'type': 'Expense'},
    
    # Maintenance Expenses
    {'code': '5350', 'name': 'Maintenance - Center', 'type': 'Expense'},
    {'code': '5360', 'name': 'Maintenance - Complex', 'type': 'Expense'},
    {'code': '5365', 'name': 'Maintenance - Computer', 'type': 'Expense'},
    {'code': '5370', 'name': 'Maintenance - Construction', 'type': 'Expense'},
    {'code': '5380', 'name': 'Maintenance - Copier', 'type': 'Expense'},
    {'code': '5385', 'name': 'Maintenance - Duplex', 'type': 'Expense'},
    {'code': '5390', 'name': 'Maintenance - Electric', 'type': 'Expense'},
    {'code': '5400', 'name': 'Maintenance - House', 'type': 'Expense'},
    {'code': '5408', 'name': 'Maintenance - Plumbing', 'type': 'Expense'},
    {'code': '5410', 'name': 'Maintenance - Landscaping/Lawn', 'type': 'Expense'},
    {'code': '5420', 'name': 'Maintenance - Other', 'type': 'Expense'},
    {'code': '5430', 'name': 'Maintenance - Restroom', 'type': 'Expense'},
    {'code': '5440', 'name': 'Maintenance - School', 'type': 'Expense'},
    {'code': '5450', 'name': 'Maintenance - Tomahawk', 'type': 'Expense'},
    {'code': '5460', 'name': 'Maintenance - Trailer', 'type': 'Expense'},
    {'code': '5470', 'name': 'Maintenance - Vehicle', 'type': 'Expense'},
    
    # Supplies
    {'code': '5790', 'name': 'Supplies - Edible', 'type': 'Expense'},
    {'code': '5800', 'name': 'Supplies - Educational', 'type': 'Expense'},
    {'code': '5810', 'name': 'Supplies - Non-Edible', 'type': 'Expense'},
    {'code': '5820', 'name': 'Supplies - Program', 'type': 'Expense'},
    {'code': '5825', 'name': 'Supplies - Office', 'type': 'Expense'},
    
    # Utilities
    {'code': '5930', 'name': 'Utilities - Electric', 'type': 'Expense'},
    {'code': '5935', 'name': 'Utilities - Gas/Fuel', 'type': 'Expense'},
    {'code': '5940', 'name': 'Utilities - Water/Sewer', 'type': 'Expense'},
    
    # Payroll & Benefits
    {'code': '6600', 'name': 'Payroll Expenses', 'type': 'Expense'},
    {'code': '6601', 'name': 'Salaries', 'type': 'Expense'},
    {'code': '6605', 'name': 'Hourly Wages', 'type': 'Expense'},
    {'code': '6625', 'name': 'Annual Bonuses', 'type': 'Expense'},
    {'code': '6700', 'name': 'Payroll Tax Expenses', 'type': 'Expense'},
    {'code': '6701', 'name': 'FICA/Medicare Taxes', 'type': 'Expense'},
    {'code': '6702', 'name': 'NC Unemployment', 'type': 'Expense'},
    {'code': '6800', 'name': 'Employee Benefits', 'type': 'Expense'},
    {'code': '6801', 'name': 'Health Insurance - Major Med', 'type': 'Expense'},
    {'code': '6802', 'name': 'Supplemental Insurance', 'type': 'Expense'},
    {'code': '6805', 'name': 'Retirement', 'type': 'Expense'},
    
    # Other Expenses
    {'code': '9510', 'name': 'Interest Expense', 'type': 'Other Expense'},
    {'code': '9550', 'name': 'Loss On Investments', 'type': 'Other Expense'},
]

# ============ COMPLETE FUND/CLASS CODES FROM CLIENT GL ============
FUND_CODES = [
    # General/Admin
    {'code': 'General Admin', 'name': 'General Admin'},
    {'code': 'Indirect Costs', 'name': 'Indirect Costs'},
    {'code': 'Fundraising', 'name': 'Fundraising'},
    
    # NAHASDA Programs
    {'code': 'NAHASDA:NAH 55IT3712260:NAHASDA #55IT3712260', 'name': 'NAHASDA #55IT3712260'},
    {'code': 'NAHASDA:NAH 55IT3712260:NAHASDA 1937 Housing Stock', 'name': 'NAHASDA 1937 Housing Stock'},
    {'code': 'NAHASDA:NAH 55IT3712260:Planning & Admin', 'name': 'NAHASDA Planning & Admin'},
    {'code': 'NAHASDA American Relief Plan', 'name': 'NAHASDA American Relief Plan'},
    
    # Boys and Girls Club Programs
    {'code': 'Boys and Girls Club', 'name': 'Boys and Girls Club'},
    {'code': 'Boys and Girls Club:BGCA - BridgeStone Tech', 'name': 'BGCA - BridgeStone Tech'},
    {'code': 'Boys and Girls Club:BGCA - FS-Overdose Prevention', 'name': 'BGCA - FS-Overdose Prevention'},
    {'code': 'Boys and Girls Club:BGCA - Juvenile Delinquency Pre', 'name': 'BGCA - Juvenile Delinquency Pre'},
    {'code': 'Boys and Girls Club:BGCA - Lilly Statewide Planning', 'name': 'BGCA - Lilly Statewide Planning'},
    {'code': 'Boys and Girls Club:BGCA - My Club Hub Adoption', 'name': 'BGCA - My Club Hub Adoption'},
    {'code': 'Boys and Girls Club:BGCA - Safety Repairs', 'name': 'BGCA - Safety Repairs'},
    {'code': 'Boys and Girls Club:BGCA - SmithSonian BookFair', 'name': 'BGCA - SmithSonian BookFair'},
    {'code': 'Boys and Girls Club:BGCA - Summer Brain Gain', 'name': 'BGCA - Summer Brain Gain'},
    {'code': 'Boys and Girls Club:BGCA - Summer Brain Gain 2', 'name': 'BGCA - Summer Brain Gain 2'},
    {'code': 'Boys and Girls Club:BGCA -LillyCultural Preservatin', 'name': 'BGCA - Lilly Cultural Preservation'},
    {'code': 'Boys and Girls Club:Boys & Girls Club Native Endo#1', 'name': 'BGC Native Endo #1'},
    {'code': 'Boys and Girls Club:Boys & Girls Club Native Endo#2', 'name': 'BGC Native Endo #2'},
    
    # State Programs
    {'code': 'NC State Recovery Funds', 'name': 'NC State Recovery Funds'},
    {'code': 'SNS', 'name': 'SNS'},
    {'code': 'NCYP FY 24-25', 'name': 'NCYP FY 24-25'},
    
    # BCBS Programs
    {'code': 'BCBS Tribal', 'name': 'BCBS Tribal'},
    {'code': 'BCBS Healthy B.H.Equity Funding', 'name': 'BCBS Healthy BH Equity Funding'},
    {'code': 'BCBS Healthy B.H.EquityFunding2', 'name': 'BCBS Healthy BH Equity Funding 2'},
    
    # Events & Cultural
    {'code': 'Pow-Wow', 'name': 'Pow-Wow'},
    {'code': 'Pow-Wow:Spring', 'name': 'Pow-Wow Spring'},
    {'code': 'Miss Haliwa Princess', 'name': 'Miss Haliwa Princess'},
    {'code': 'Tiny, Little, Jr Miss Pageant', 'name': 'Tiny, Little, Jr Miss Pageant'},
    {'code': 'July 4th', 'name': 'July 4th'},
    {'code': 'Awards Banquet', 'name': 'Awards Banquet'},
    {'code': 'Cultural Prevention/ Cultural C', 'name': 'Cultural Prevention'},
    {'code': 'Red Earth Cultural Group', 'name': 'Red Earth Cultural Group'},
    
    # Properties & Community
    {'code': 'Tomahawk Meadows', 'name': 'Tomahawk Meadows'},
    {'code': 'Rental Properties', 'name': 'Rental Properties'},
    {'code': 'Veterans Memorial', 'name': 'Veterans Memorial'},
    
    # Grants & Foundations
    {'code': 'Kate B. Reynolds', 'name': 'Kate B. Reynolds'},
    {'code': 'Farrington Foundation Grant', 'name': 'Farrington Foundation Grant'},
    {'code': 'USDA - High Tunnel', 'name': 'USDA - High Tunnel'},
    {'code': 'SAMHSA Three Star Gov\'t Solutio', 'name': 'SAMHSA Three Star Govt Solution'},
    {'code': 'TAPS:TAPS 20-21', 'name': 'TAPS 20-21'},
    
    # Tribal Programs
    {'code': 'Federal Recognition', 'name': 'Federal Recognition'},
    {'code': 'Economic Development', 'name': 'Economic Development'},
    {'code': 'Health and Human Services', 'name': 'Health and Human Services'},
]

# Vendors from GL
VENDORS = [
    {'name': 'Delta Dental of North Carolina', 'gl_code': '24060', 'fund': 'General Admin', 'type': 'Vendor'},
    {'name': 'Quality Office Equipment Inc', 'gl_code': '5380', 'fund': 'SNS', 'type': 'Vendor'},
    {'name': 'PROSHRED of Raleigh', 'gl_code': '5810', 'fund': 'Indirect Costs', 'type': 'Vendor'},
    {'name': 'County Line Rental', 'gl_code': '5086', 'fund': 'NAHASDA:NAH 55IT3712260:NAHASDA #55IT3712260', 'type': 'Vendor'},
    {'name': 'GFL Environmental', 'gl_code': '5950', 'fund': 'Indirect Costs', 'type': 'Vendor'},
    {'name': 'Regina Mills', 'gl_code': '5200', 'fund': 'SNS', 'type': 'Vendor'},
    {'name': 'Dorothy Lynch', 'gl_code': '5810', 'fund': 'Miss Haliwa Princess', 'type': 'Vendor'},
    {'name': 'R - Sharon Berrun', 'gl_code': '5810', 'fund': 'Miss Haliwa Princess', 'type': 'Vendor'},
    {'name': 'Sheena Richardson', 'gl_code': '5505', 'fund': 'Miss Haliwa Princess', 'type': 'Vendor'},
    {'name': 'PNC Bank - 0580', 'gl_code': '5060', 'fund': 'General Admin', 'type': 'Vendor'},
    {'name': 'ADP Payroll', 'gl_code': '6600', 'fund': 'General Admin', 'type': 'Vendor'},
    {'name': 'IRS / EFTPS', 'gl_code': '6700', 'fund': 'General Admin', 'type': 'Vendor'},
]

# Customers from GL
CUSTOMERS = [
    {'name': 'NAHASDA 55IT3712260', 'gl_code': '3001', 'fund': 'NAHASDA:NAH 55IT3712260:NAHASDA #55IT3712260', 'type': 'Customer'},
    {'name': 'NAHASDA 1937 Housing Stock', 'gl_code': '3001', 'fund': 'NAHASDA:NAH 55IT3712260:NAHASDA 1937 Housing Stock', 'type': 'Customer'},
    {'name': 'NAHASDA ARP', 'gl_code': '3001', 'fund': 'NAHASDA American Relief Plan', 'type': 'Customer'},
    {'name': 'Boys and Girls Club', 'gl_code': '3001', 'fund': 'Boys and Girls Club', 'type': 'Customer'},
    {'name': 'BGCA - Juvenile Delinquency Pre', 'gl_code': '3001', 'fund': 'Boys and Girls Club:BGCA - Juvenile Delinquency Pre', 'type': 'Customer'},
    {'name': 'BGCA - Lilly Statewide Planning', 'gl_code': '3001', 'fund': 'Boys and Girls Club:BGCA - Lilly Statewide Planning', 'type': 'Customer'},
    {'name': 'BGCA - SmithSonian BookFair', 'gl_code': '3001', 'fund': 'Boys and Girls Club:BGCA - SmithSonian BookFair', 'type': 'Customer'},
    {'name': 'BGCA - FS-Overdose Prevention', 'gl_code': '3001', 'fund': 'Boys and Girls Club:BGCA - FS-Overdose Prevention', 'type': 'Customer'},
    {'name': 'BGCA - BridgeStone Tech', 'gl_code': '3001', 'fund': 'Boys and Girls Club:BGCA - BridgeStone Tech', 'type': 'Customer'},
    {'name': 'BGCA - Summer Brain Gain', 'gl_code': '3001', 'fund': 'Boys and Girls Club:BGCA - Summer Brain Gain', 'type': 'Customer'},
    {'name': 'BGCA - Summer Brain Gain 2', 'gl_code': '3001', 'fund': 'Boys and Girls Club:BGCA - Summer Brain Gain 2', 'type': 'Customer'},
    {'name': 'BGCA - Safety Repairs', 'gl_code': '3001', 'fund': 'Boys and Girls Club:BGCA - Safety Repairs', 'type': 'Customer'},
    {'name': 'BGCA - My Club Hub Adoption', 'gl_code': '3001', 'fund': 'Boys and Girls Club:BGCA - My Club Hub Adoption', 'type': 'Customer'},
    {'name': 'BGCA - Lilly Cultural Preservatin', 'gl_code': '3001', 'fund': 'Boys and Girls Club:BGCA -LillyCultural Preservatin', 'type': 'Customer'},
    {'name': 'BGC Native Endo #1', 'gl_code': '3001', 'fund': 'Boys and Girls Club:Boys & Girls Club Native Endo#1', 'type': 'Customer'},
    {'name': 'BGC Native Endo #2', 'gl_code': '3001', 'fund': 'Boys and Girls Club:Boys & Girls Club Native Endo#2', 'type': 'Customer'},
    {'name': 'BCBS Tribal', 'gl_code': '4080', 'fund': 'BCBS Tribal', 'type': 'Customer'},
    {'name': 'BCBS Healthy BH Equity Funding', 'gl_code': '3002', 'fund': 'BCBS Healthy B.H.Equity Funding', 'type': 'Customer'},
    {'name': 'BCBS Healthy BH Equity Funding 2', 'gl_code': '3002', 'fund': 'BCBS Healthy B.H.EquityFunding2', 'type': 'Customer'},
    {'name': 'NC State Recovery Funds', 'gl_code': '3002', 'fund': 'NC State Recovery Funds', 'type': 'Customer'},
    {'name': 'Veterans Memorial', 'gl_code': '4100', 'fund': 'Veterans Memorial', 'type': 'Customer'},
    {'name': 'SNS', 'gl_code': '3002', 'fund': 'SNS', 'type': 'Customer'},
    {'name': 'TAPS 20-21', 'gl_code': '3002', 'fund': 'TAPS:TAPS 20-21', 'type': 'Customer'},
    {'name': 'Farrington Foundation Grant', 'gl_code': '4225', 'fund': 'Farrington Foundation Grant', 'type': 'Customer'},
    {'name': 'SAMHSA Three Star Govt Solution', 'gl_code': '3001', 'fund': 'SAMHSA Three Star Gov\'t Solutio', 'type': 'Customer'},
    {'name': 'NCYP FY 24-25', 'gl_code': '3002', 'fund': 'NCYP FY 24-25', 'type': 'Customer'},
    {'name': 'HUD', 'gl_code': '3001', 'fund': 'NAHASDA:NAH 55IT3712260:NAHASDA #55IT3712260', 'type': 'Customer'},
    {'name': 'Kate B. Reynolds', 'gl_code': '4080', 'fund': 'Kate B. Reynolds', 'type': 'Customer'},
    {'name': 'USDA - High Tunnel', 'gl_code': '3001', 'fund': 'USDA - High Tunnel', 'type': 'Customer'},
    {'name': 'Red Earth Cultural Group', 'gl_code': '4080', 'fund': 'Red Earth Cultural Group', 'type': 'Customer'},
    {'name': 'Federal Recognition', 'gl_code': '3002', 'fund': 'Federal Recognition', 'type': 'Customer'},
    {'name': 'Tomahawk Meadows', 'gl_code': '4380', 'fund': 'Tomahawk Meadows', 'type': 'Customer'},
    {'name': 'Pow-Wow Spring', 'gl_code': '4100', 'fund': 'Pow-Wow:Spring', 'type': 'Customer'},
    {'name': 'July 4th Event', 'gl_code': '4100', 'fund': 'July 4th', 'type': 'Customer'},
]

def get_all_vendors():
    """Get all vendors from MongoDB or fallback to static list"""
    db = get_db()
    if db is not None:
        try:
            vendors = list(db.vendors.find({}, {'_id': 0}))
            if vendors:
                return vendors
        except Exception as e:
            print(f"Error loading vendors from MongoDB: {e}")
    return VENDORS

def get_all_customers():
    """Get all customers from MongoDB or fallback to static list"""
    db = get_db()
    if db is not None:
        try:
            customers = list(db.customers.find({}, {'_id': 0}))
            if customers:
                return customers
        except Exception as e:
            print(f"Error loading customers from MongoDB: {e}")
    return CUSTOMERS

SUPPORTED_EXTENSIONS = {'.pdf', '.xlsx', '.xls', '.csv'}

def allowed_file(filename):
    return os.path.splitext(filename)[1].lower() in SUPPORTED_EXTENSIONS

def get_user_session_id():
    """Get or create unique session ID for current user"""
    import uuid
    if 'user_session_id' not in session:
        session['user_session_id'] = str(uuid.uuid4())
    return session['user_session_id']

def get_user_session_data():
    """Get session data for current user from MongoDB or fallback to session cookie"""
    session_id = get_user_session_id()
    db = get_db()

    if db is not None:
        try:
            session_doc = db.user_sessions.find_one({'session_id': session_id})
            if session_doc:
                return session_doc.get('data', {})
        except Exception as e:
            print(f"Error loading user session: {e}")

    # Fallback to Flask session (less secure, for compatibility)
    return session.get('session_data', {'transactions': [], 'classified': [], 'audit_trail': [], 'output_files': []})

def save_user_session_data(data):
    """Save session data for current user to MongoDB"""
    session_id = get_user_session_id()
    db = get_db()

    if db is not None:
        try:
            db.user_sessions.update_one(
                {'session_id': session_id},
                {
                    '$set': {
                        'session_id': session_id,
                        'data': data,
                        'updated_at': datetime.now(),
                        'user_agent': request.headers.get('User-Agent', ''),
                        'ip_address': request.remote_addr
                    }
                },
                upsert=True
            )
            return True
        except Exception as e:
            print(f"Error saving user session: {e}")

    # Fallback to Flask session
    session['session_data'] = data
    return True

def generate_transaction_hash(txn):
    """Generate unique hash for duplicate detection"""
    import hashlib

    check_number = txn.get('check_number', '')
    description = txn.get('description', '').upper()
    is_check = bool(check_number) or 'CHECK' in description
    is_deposit = txn.get('is_deposit', False) or txn.get('module') == 'CR' or 'DEPOSIT' in description

    if is_check and check_number:
        # For checks, use check number as primary identifier (checks are unique)
        key = f"CHECK|{check_number}|{txn.get('amount', 0)}"
    elif is_deposit:
        # For deposits, include a unique index to allow legitimate duplicates
        # Banks legitimately process multiple deposits of same amount on same day
        # Use reference_number if available, otherwise use parsed_by + position hint
        ref_num = txn.get('reference_number', '') or txn.get('deposit_slip_number', '')
        if ref_num:
            key = f"{txn.get('date', '')}|{txn.get('amount', 0)}|DEPOSIT|{ref_num}"
        else:
            # For deposits without reference numbers, create unique key using position
            # This prevents flagging legitimate same-day, same-amount deposits
            position = txn.get('_position', id(txn))  # Use object id as fallback
            key = f"{txn.get('date', '')}|{txn.get('amount', 0)}|DEPOSIT|pos_{position}"
    else:
        # For other transactions (fees, etc.), use full details
        key = f"{txn.get('date', '')}|{txn.get('amount', 0)}|{description}|{check_number}"

    return hashlib.sha256(key.encode()).hexdigest()[:16]

def check_for_duplicates(transactions):
    """
    Check for duplicate transactions in MongoDB and within current batch.

    Duplicate detection rules:
    - CHECKS: Duplicate if same check number (checks are unique identifiers)
    - DEPOSITS: NOT duplicates just because same date/amount (banks process multiple deposits daily)
    - FEES/OTHER: Duplicate if same date + amount + description
    """
    db = get_db()
    duplicates_found = []
    seen_in_batch = {}  # Changed to dict to track check numbers specifically
    seen_check_numbers = set()

    for i, txn in enumerate(transactions):
        # Add position hint for hash generation
        txn['_position'] = i

        # Generate hash
        txn_hash = generate_transaction_hash(txn)
        txn['txn_hash'] = txn_hash

        check_number = txn.get('check_number', '')
        description = txn.get('description', '').upper()
        is_check = bool(check_number) or 'CHECK' in description
        is_deposit = txn.get('is_deposit', False) or txn.get('module') == 'CR' or 'DEPOSIT' in description

        # Special handling for checks - check numbers must be unique
        if is_check and check_number:
            if check_number in seen_check_numbers:
                txn['is_duplicate'] = True
                txn['duplicate_source'] = 'current_batch'
                txn['duplicate_reason'] = f'Duplicate check number: {check_number}'
                duplicates_found.append(i)
                continue
            else:
                seen_check_numbers.add(check_number)

        # For deposits, do NOT flag as duplicate within same batch
        # Multiple deposits of same amount on same day are legitimate
        if is_deposit:
            txn['is_duplicate'] = False
            # Still check MongoDB for cross-batch duplicates using reference numbers
            if db is not None and txn.get('reference_number'):
                try:
                    existing = db.transactions.find_one({
                        'reference_number': txn['reference_number'],
                        'date': txn.get('date'),
                        'amount': txn.get('amount')
                    })
                    if existing:
                        txn['is_duplicate'] = True
                        txn['duplicate_source'] = 'database'
                        txn['duplicate_batch_id'] = existing.get('batch_id')
                        duplicates_found.append(i)
                except Exception as e:
                    print(f"Error checking deposit duplicates: {e}")
            continue

        # For non-deposit, non-check transactions (fees, etc.)
        if txn_hash in seen_in_batch:
            txn['is_duplicate'] = True
            txn['duplicate_source'] = 'current_batch'
            duplicates_found.append(i)
        else:
            seen_in_batch[txn_hash] = i

            # Check in MongoDB
            if db is not None:
                try:
                    existing = db.transactions.find_one({'txn_hash': txn_hash})
                    if existing:
                        txn['is_duplicate'] = True
                        txn['duplicate_source'] = 'database'
                        txn['duplicate_batch_id'] = existing.get('batch_id')
                        duplicates_found.append(i)
                    else:
                        txn['is_duplicate'] = False
                except Exception as e:
                    print(f"Error checking duplicates: {e}")
                    txn['is_duplicate'] = False
            else:
                txn['is_duplicate'] = False

    # Clean up position hints
    for txn in transactions:
        txn.pop('_position', None)

    return duplicates_found

@app.route('/')
def index():
    return render_template_string(INDEX_TEMPLATE)

@app.route('/clear_session')
def clear_session():
    """Clear user session data to force fresh upload"""
    session_id = get_user_session_id()
    db = get_db()

    if db is not None:
        try:
            db.user_sessions.delete_one({'session_id': session_id})
            print(f"[DEBUG] Cleared MongoDB session: {session_id}")
        except Exception as e:
            print(f"Error clearing session: {e}")

    # Clear Flask session too
    session.clear()
    flash('Session cleared. Upload a new file.', 'success')
    return redirect(url_for('index'))

@app.route('/upload', methods=['POST'])
def upload():
    # Support both single file ('file') and multiple files ('files')
    files = request.files.getlist('files')
    if not files or (len(files) == 1 and files[0].filename == ''):
        # Fallback to single file upload for backwards compatibility
        if 'file' in request.files:
            single_file = request.files['file']
            if single_file.filename != '':
                files = [single_file]
            else:
                flash('No file selected', 'error')
                return redirect(url_for('index'))
        else:
            flash('No file selected', 'error')
            return redirect(url_for('index'))

    # Filter valid files
    valid_files = [f for f in files if f.filename != '' and allowed_file(f.filename)]
    if not valid_files:
        flash('No valid files selected. Supported formats: PDF, Excel, CSV', 'error')
        return redirect(url_for('index'))

    try:
        import tempfile
        import sys

        all_transactions = []
        all_filenames = []
        combined_metadata = {
            'files_processed': [],
            'ocr_used': False,
            'bank_name': None,
            'warnings': []
        }

        print(f"", flush=True)
        print(f"=" * 60, flush=True)
        print(f"[DEBUG] *** PROCESSING {len(valid_files)} FILE(S) ***", flush=True)
        print(f"=" * 60, flush=True)

        for file_idx, file in enumerate(valid_files):
            filename = secure_filename(file.filename)
            all_filenames.append(filename)

            print(f"", flush=True)
            print(f"[DEBUG] Processing file {file_idx + 1}/{len(valid_files)}: {filename}", flush=True)

            # Process file in-memory using a temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(filename)[1]) as tmp_file:
                file.save(tmp_file.name)
                tmp_filepath = tmp_file.name

            # Use SmartParser - template-based with AI fallback
            parser = UniversalParser(use_llm=False)
            print(f"[DEBUG] Parser type: {type(parser.smart_parser).__name__ if hasattr(parser, 'smart_parser') else 'NO SMART PARSER!'}", flush=True)
            print(f"[DEBUG] File exists: {os.path.exists(tmp_filepath)}", flush=True)
            print(f"[DEBUG] File size: {os.path.getsize(tmp_filepath)} bytes", flush=True)
            sys.stdout.flush()

            try:
                transactions = parser.parse(tmp_filepath)
                print(f"[DEBUG] Found {len(transactions)} transactions in {filename}", flush=True)

                # Add source file info to each transaction
                for txn in transactions:
                    txn['source_file'] = filename

                # DEBUG: Show each transaction amount
                for i, txn in enumerate(transactions):
                    print(f"[DEBUG] Txn {i+1}: amount={txn.get('amount')}, desc={txn.get('description', '')[:30]}, module={txn.get('module')}", flush=True)

                # Capture parsing metadata for validation warnings
                parsing_metadata = parser.get_parsing_metadata()
                print(f"[DEBUG] Parsing metadata: {parsing_metadata}", flush=True)

                # Merge metadata
                combined_metadata['files_processed'].append({
                    'filename': filename,
                    'transaction_count': len(transactions),
                    'bank_name': parsing_metadata.get('bank_name'),
                    'ocr_used': parsing_metadata.get('ocr_used', False)
                })
                if parsing_metadata.get('ocr_used'):
                    combined_metadata['ocr_used'] = True
                if parsing_metadata.get('bank_name') and not combined_metadata['bank_name']:
                    combined_metadata['bank_name'] = parsing_metadata.get('bank_name')
                if parsing_metadata.get('warnings'):
                    combined_metadata['warnings'].extend(parsing_metadata.get('warnings', []))

                all_transactions.extend(transactions)

            except Exception as parse_error:
                import traceback
                print(f"[ERROR] Parse failed for {filename}: {parse_error}", flush=True)
                traceback.print_exc()
                combined_metadata['warnings'].append(f"Failed to parse {filename}: {str(parse_error)}")

            # Clean up temporary file
            try:
                os.unlink(tmp_filepath)
            except:
                pass

        if not all_transactions:
            flash('No transactions found in any of the uploaded files.', 'error')
            return redirect(url_for('index'))

        print(f"", flush=True)
        print(f"[DEBUG] Total transactions from all files: {len(all_transactions)}", flush=True)

        # Check for duplicates across all files
        duplicates_found = check_for_duplicates(all_transactions)

        # Validate transactions
        validation_warnings = []
        for i, txn in enumerate(all_transactions):
            # Check for zero amounts
            if abs(txn.get('amount', 0)) < 0.01:
                txn['needs_review'] = True
                txn['review_reason'] = 'Zero or near-zero amount'
                validation_warnings.append(f"Line {i+1}: Zero amount transaction")

            # Check for future dates
            try:
                txn_date = datetime.strptime(str(txn.get('date')), '%Y-%m-%d') if isinstance(txn.get('date'), str) else txn.get('date')
                if txn_date and txn_date > datetime.now():
                    txn['needs_review'] = True
                    txn['review_reason'] = txn.get('review_reason', '') + ' | Future-dated'
                    validation_warnings.append(f"Line {i+1}: Future-dated transaction")
            except:
                pass

        classifier = ClassificationEngine()
        classified = classifier.classify_batch(all_transactions)

        # DEBUG: Show classified amounts
        print(f"[DEBUG] After classification:", flush=True)
        for i, txn in enumerate(classified):
            print(f"[DEBUG] Classified {i+1}: amount={txn.get('amount')}, module={txn.get('module')}", flush=True)

        # Calculate totals by module (not just by sign) to avoid double-counting
        by_module_temp = {'CR': [], 'CD': [], 'JV': [], 'UNKNOWN': []}
        for t in classified:
            by_module_temp[t.get('module', 'UNKNOWN')].append(t)
        total_cr = sum(abs(t.get('amount', 0)) for t in by_module_temp['CR'])
        total_cd = sum(abs(t.get('amount', 0)) for t in by_module_temp['CD'])
        total_jv = sum(abs(t.get('amount', 0)) for t in by_module_temp['JV'])
        print(f"[DEBUG] Summary - CR: ${total_cr:.2f}, CD: ${total_cd:.2f}, JV: ${total_jv:.2f}", flush=True)

        # Create combined filename for display
        combined_filename = ', '.join(all_filenames) if len(all_filenames) <= 3 else f"{all_filenames[0]} + {len(all_filenames) - 1} more files"

        # FIXED: Use per-user session instead of global variable
        session_data = {
            'transactions': all_transactions,
            'classified': classified,
            'audit_trail': [],
            'output_files': [],
            'uploaded_at': datetime.now().isoformat(),
            'filename': combined_filename,
            'filenames': all_filenames,  # Store all filenames
            'parsing_metadata': combined_metadata  # Store combined metadata
        }
        print(f"[DEBUG] SAVING SESSION: {len(all_transactions)} raw, {len(classified)} classified", flush=True)
        save_result = save_user_session_data(session_data)
        print(f"[DEBUG] Save result: {save_result}", flush=True)

        # Build flash message with warnings
        if len(valid_files) > 1:
            messages = [f'Processed {len(valid_files)} files with {len(all_transactions)} total transactions']
        else:
            messages = [f'Parsed {len(all_transactions)} transactions']

        if duplicates_found:
            messages.append(f'[!] {len(duplicates_found)} potential duplicates found (flagged for review)')
        if validation_warnings:
            messages.append(f'[!] {len(validation_warnings)} transactions need review')

        for msg in messages:
            flash(msg, 'warning' if '[!]' in msg else 'success')

        return redirect(url_for('review'))
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"UPLOAD ERROR: {error_trace}")  # Print to console for debugging
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/review')
def review():
    session_data = get_user_session_data()
    classified = session_data.get('classified', [])
    if not classified:
        flash('No transactions. Upload a file first.', 'warning')
        return redirect(url_for('index'))

    by_module = {'CR': [], 'CD': [], 'JV': [], 'UNKNOWN': []}
    for txn in classified:
        by_module[txn.get('module', 'UNKNOWN')].append(txn)

    # Calculate totals BY MODULE to avoid double-counting
    # CR total: sum of positive amounts for CR transactions
    total_cr = sum(abs(t.get('amount', 0)) for t in by_module['CR'])
    # CD total: sum of negative amounts for CD transactions ONLY (excludes JV fees)
    total_cd = sum(abs(t.get('amount', 0)) for t in by_module['CD'])
    # JV total: sum of all JV transaction amounts (fees/interest)
    total_jv = sum(abs(t.get('amount', 0)) for t in by_module['JV'])

    # Net cash flow = deposits - withdrawals (CD only, not JV)
    # JV transactions (bank fees, interest) are internal adjustments, not cash flow
    net_cash_flow = total_cr - total_cd

    # Calculate year-segregated totals for audit trail
    by_year = {}
    for txn in classified:
        # Extract year from date (format: MM/DD/YYYY)
        date_str = txn.get('date', '')
        try:
            if '/' in date_str:
                parts = date_str.split('/')
                if len(parts) == 3:
                    year = int(parts[2])
                else:
                    year = 'Unknown'
            else:
                year = 'Unknown'
        except:
            year = 'Unknown'

        if year not in by_year:
            by_year[year] = {'deposits': 0, 'withdrawals': 0, 'count': 0, 'transactions': []}

        by_year[year]['count'] += 1
        by_year[year]['transactions'].append(txn)

        if txn.get('module') == 'CR':
            by_year[year]['deposits'] += abs(txn.get('amount', 0))
        elif txn.get('module') == 'CD':
            by_year[year]['withdrawals'] += abs(txn.get('amount', 0))

    # Get parsing metadata for validation warnings display
    parsing_metadata = session_data.get('parsing_metadata', {})

    # Create response with no-cache headers to prevent browser caching
    response = make_response(render_template_string(REVIEW_TEMPLATE,
        transactions=classified, by_module=by_module,
        summary={'total_credits': total_cr, 'total_debits': total_cd, 'total_jv': total_jv, 'balance': net_cash_flow},
        by_year=by_year,
        gl_codes=GL_CODES, fund_codes=FUND_CODES,
        vendors=get_all_vendors(), customers=get_all_customers(),
        audit_trail=session_data.get('audit_trail', []),
        parsing_metadata=parsing_metadata))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/update_transaction', methods=['POST'])
def update_transaction():
    try:
        session_data = get_user_session_data()
        data = request.get_json()
        idx = data.get('index')
        if idx is None or idx >= len(session_data['classified']):
            return jsonify({'status': 'error', 'message': 'Invalid index'})

        txn = session_data['classified'][idx]
        changes = []

        for field in ['module', 'gl_code', 'fund_code', 'payee']:
            if data.get(field) and data[field] != txn.get(field):
                changes.append(f"{field}: {txn.get(field)} -> {data[field]}")
                txn[field] = data[field]

        if changes:
            if 'audit_trail' not in session_data:
                session_data['audit_trail'] = []
            session_data['audit_trail'].append({
                'timestamp': datetime.now().strftime('%H:%M:%S'),
                'transaction': txn.get('description', '')[:25],
                'changes': changes,
                'action': 'Edit'
            })
            txn['confidence_level'] = 'manual'

        # Save updated session data
        save_user_session_data(session_data)

        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/add_transaction', methods=['POST'])
def add_transaction():
    """Add a new manual transaction"""
    try:
        session_data = get_user_session_data()
        data = request.get_json()

        # Validate required fields
        date = data.get('date')
        amount = data.get('amount')
        description = data.get('description', '').strip()

        if not date or amount is None or not description:
            return jsonify({'status': 'error', 'message': 'Date, amount, and description are required'})

        # Create new transaction
        new_txn = {
            'date': date,
            'amount': float(amount),
            'description': description,
            'module': data.get('module', 'CR' if float(amount) > 0 else 'CD'),
            'gl_code': data.get('gl_code', ''),
            'fund_code': data.get('fund_code', ''),
            'confidence_level': 'manual',
            'confidence_score': 100,
            'is_deposit': float(amount) > 0,
            'manually_added': True
        }

        # Add to classified transactions
        if 'classified' not in session_data:
            session_data['classified'] = []
        session_data['classified'].append(new_txn)

        # Add audit trail entry
        if 'audit_trail' not in session_data:
            session_data['audit_trail'] = []
        session_data['audit_trail'].append({
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'transaction': f'ADDED: {description[:25]}',
            'changes': [f'amount: ${abs(float(amount)):,.2f}', f'module: {new_txn["module"]}'],
            'action': 'Add'
        })

        save_user_session_data(session_data)
        return jsonify({'status': 'success', 'message': f'Transaction "{description[:30]}" added'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/delete_transaction', methods=['POST'])
def delete_transaction():
    """Delete a transaction by index"""
    try:
        session_data = get_user_session_data()
        data = request.get_json()
        idx = data.get('index')

        if idx is None or idx >= len(session_data.get('classified', [])):
            return jsonify({'status': 'error', 'message': 'Invalid index'})

        # Get transaction info before deleting
        txn = session_data['classified'][idx]
        desc = txn.get('description', 'Unknown')[:25]
        amount = txn.get('amount', 0)

        # Remove transaction
        session_data['classified'].pop(idx)

        # Add audit trail entry
        if 'audit_trail' not in session_data:
            session_data['audit_trail'] = []
        session_data['audit_trail'].append({
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'transaction': f'DELETED: {desc}',
            'changes': [f'amount: ${abs(amount):,.2f}'],
            'action': 'Delete'
        })

        save_user_session_data(session_data)
        return jsonify({'status': 'success', 'message': f'Transaction deleted'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/bulk_update', methods=['POST'])
def bulk_update():
    try:
        session_data = get_user_session_data()
        data = request.get_json()
        indices = data.get('indices', [])
        updates = data.get('updates', {})

        for idx in indices:
            if idx < len(session_data['classified']):
                txn = session_data['classified'][idx]
                for k, v in updates.items():
                    if v: txn[k] = v
                txn['confidence_level'] = 'manual'

        if 'audit_trail' not in session_data:
            session_data['audit_trail'] = []
        session_data['audit_trail'].append({
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'transaction': f'Bulk ({len(indices)} items)',
            'changes': [f"{k}={v}" for k,v in updates.items() if v],
            'action': 'Bulk'
        })

        # Save updated session data
        save_user_session_data(session_data)

        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/undo_last', methods=['POST'])
def undo_last():
    session_data = get_user_session_data()
    if session_data.get('audit_trail'):
        entry = session_data['audit_trail'].pop()
        save_user_session_data(session_data)
        return jsonify({'status': 'success', 'message': f"Undone: {entry['transaction']}"})
    return jsonify({'status': 'error', 'message': 'Nothing to undo'})

@app.route('/add_vendor', methods=['POST'])
def add_vendor():
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        gl_code = data.get('gl_code', '').strip()
        fund = data.get('fund', 'General').strip()
        
        if not name:
            return jsonify({'status': 'error', 'message': 'Vendor name is required'})
        
        all_vendors = get_all_vendors()
        if any(v['name'].lower() == name.lower() for v in all_vendors):
            return jsonify({'status': 'error', 'message': 'Vendor already exists'})

        # Save to MongoDB if available
        db = get_db()
        if db is not None:
            try:
                db.vendors.insert_one({'name': name, 'gl_code': gl_code, 'fund': fund, 'type': 'Vendor', 'created_at': datetime.now()})
            except Exception as e:
                print(f"Error saving vendor to MongoDB: {e}")

        return jsonify({'status': 'success', 'message': f'Vendor "{name}" added'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/add_customer', methods=['POST'])
def add_customer():
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        gl_code = data.get('gl_code', '').strip()
        fund = data.get('fund', 'General').strip()
        
        if not name:
            return jsonify({'status': 'error', 'message': 'Customer name is required'})
        
        all_customers = get_all_customers()
        if any(c['name'].lower() == name.lower() for c in all_customers):
            return jsonify({'status': 'error', 'message': 'Customer already exists'})

        # Save to MongoDB if available
        db = get_db()
        if db is not None:
            try:
                db.customers.insert_one({'name': name, 'gl_code': gl_code, 'fund': fund, 'type': 'Customer', 'created_at': datetime.now()})
            except Exception as e:
                print(f"Error saving customer to MongoDB: {e}")

        return jsonify({'status': 'success', 'message': f'Customer "{name}" added'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/process', methods=['POST'])
def process():
    try:
        session_data = get_user_session_data()
        classified = session_data.get('classified', [])
        if not classified:
            flash('No transactions', 'error')
            return redirect(url_for('review'))
        
        router = ModuleRouter()
        routed_by_module = router.route_batch(classified)
        
        # Generate files in-memory and store in MongoDB
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        batch_id = f"BATCH_{timestamp}"
        files_list = []
        
        # Get MongoDB connection
        db = get_db()
        
        # Generate Excel file for each module with transactions
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        module_names = {
            'CR': 'Cash_Receipts',
            'CD': 'Cash_Disbursements', 
            'JV': 'Journal_Voucher'
        }
        
        # Use routed_by_module directly - already grouped by module
        for module in ['CR', 'CD', 'JV']:
            module_entries = routed_by_module.get(module, [])
            if module_entries:
                filename = f"{module_names[module]}_{timestamp}.xlsx"
                
                # Create workbook in memory
                buffer = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = module_names[module]
                
                # Add headers
                headers = ['Date', 'Description', 'Amount', 'GL Code', 'Fund Code', 'Module', 'Payee', 'Reference']
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Add data rows - handle both dict and other formats
                for row_idx, entry in enumerate(module_entries, 2):
                    if isinstance(entry, dict):
                        ws.cell(row=row_idx, column=1, value=entry.get('date', ''))
                        ws.cell(row=row_idx, column=2, value=entry.get('description', ''))
                        ws.cell(row=row_idx, column=3, value=entry.get('amount', 0))
                        ws.cell(row=row_idx, column=4, value=entry.get('gl_code', ''))
                        ws.cell(row=row_idx, column=5, value=entry.get('fund_code', ''))
                        ws.cell(row=row_idx, column=6, value=module)
                        ws.cell(row=row_idx, column=7, value=entry.get('payee', entry.get('vendor', '')))
                        ws.cell(row=row_idx, column=8, value=entry.get('reference', entry.get('check_number', '')))
                    else:
                        # Handle non-dict entries (convert to string)
                        ws.cell(row=row_idx, column=2, value=str(entry))
                        ws.cell(row=row_idx, column=6, value=module)
                
                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
                
                wb.save(buffer)
                buffer.seek(0)
                file_data = base64.b64encode(buffer.getvalue()).decode('utf-8')
                
                # Store in MongoDB if available
                file_doc = {
                    'filename': filename,
                    'batch_id': batch_id,
                    'module': module,
                    'description': f'{module_names[module]} Import File ({len(module_entries)} entries)',
                    'file_data': file_data,
                    'file_size': len(buffer.getvalue()),
                    'entry_count': len(module_entries),
                    'created_at': datetime.now(),
                    'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                }
                
                if db is not None:
                    result = db.output_files.insert_one(file_doc)
                    file_doc['_id'] = result.inserted_id
                
                files_list.append({
                    'filename': filename, 
                    'description': f'{module_names[module]} Import File ({len(module_entries)} entries)',
                    'file_id': str(file_doc.get('_id', filename))
                })
        
        # Generate Unidentified file if any
        unidentified = routed_by_module.get('UNIDENTIFIED', [])
        if unidentified:
            filename = f"Unidentified_{timestamp}.xlsx"
            buffer = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = 'Unidentified'
            
            headers = ['Date', 'Description', 'Amount', 'Suggested GL', 'Suggested Fund', 'Confidence', 'Notes']
            header_fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            for row_idx, txn in enumerate(unidentified, 2):
                if isinstance(txn, dict):
                    ws.cell(row=row_idx, column=1, value=txn.get('date', ''))
                    ws.cell(row=row_idx, column=2, value=txn.get('description', ''))
                    ws.cell(row=row_idx, column=3, value=txn.get('amount', 0))
                    ws.cell(row=row_idx, column=4, value=txn.get('gl_code', ''))
                    ws.cell(row=row_idx, column=5, value=txn.get('fund_code', ''))
                    ws.cell(row=row_idx, column=6, value=txn.get('confidence_level', 'low'))
                    ws.cell(row=row_idx, column=7, value='Needs manual review')
                else:
                    ws.cell(row=row_idx, column=2, value=str(txn))
                    ws.cell(row=row_idx, column=7, value='Needs manual review')
            
            for col in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
            
            wb.save(buffer)
            buffer.seek(0)
            file_data = base64.b64encode(buffer.getvalue()).decode('utf-8')
            
            file_doc = {
                'filename': filename,
                'batch_id': batch_id,
                'module': 'UNIDENTIFIED',
                'description': f'Unidentified Transactions ({len(unidentified)} entries) - Needs Review',
                'file_data': file_data,
                'file_size': len(buffer.getvalue()),
                'entry_count': len(unidentified),
                'created_at': datetime.now(),
                'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            if db is not None:
                result = db.output_files.insert_one(file_doc)
                file_doc['_id'] = result.inserted_id
            
            files_list.append({
                'filename': filename, 
                'description': f'Unidentified Transactions ({len(unidentified)} entries) - Needs Review',
                'file_id': str(file_doc.get('_id', filename))
            })
        
        # Generate Summary Report
        filename = f"Processing_Summary_{timestamp}.xlsx"
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'
        
        # Summary header
        ws.cell(row=1, column=1, value='Bank Transaction Processing Summary')
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f'Generated: {datetime.now().strftime("%m/%d/%Y %H:%M:%S")}')
        
        ws.cell(row=4, column=1, value='Module')
        ws.cell(row=4, column=2, value='Count')
        ws.cell(row=4, column=3, value='Total Amount')
        
        row = 5
        total_count = 0
        total_amount = 0
        for module in ['CR', 'CD', 'JV', 'UNIDENTIFIED']:
            module_txns = routed_by_module.get(module, [])
            count = len(module_txns)
            # Safely calculate amount - handle both dict and non-dict entries
            amount = 0
            for t in module_txns:
                if isinstance(t, dict):
                    amount += abs(t.get('amount', 0) or 0)
            ws.cell(row=row, column=1, value=module)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=amount)
            total_count += count
            total_amount += amount
            row += 1
        
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_count)
        ws.cell(row=row, column=3, value=total_amount)
        
        wb.save(buffer)
        buffer.seek(0)
        file_data = base64.b64encode(buffer.getvalue()).decode('utf-8')
        
        file_doc = {
            'filename': filename,
            'batch_id': batch_id,
            'module': 'SUMMARY',
            'description': 'Processing Summary Report',
            'file_data': file_data,
            'file_size': len(buffer.getvalue()),
            'entry_count': total_count,
            'created_at': datetime.now(),
            'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
        if db is not None:
            result = db.output_files.insert_one(file_doc)
            file_doc['_id'] = result.inserted_id
        
        files_list.append({
            'filename': filename, 
            'description': 'Processing Summary Report',
            'file_id': str(file_doc.get('_id', filename))
        })
        
        # Store batch_id in session for results page
        session_data['output_files'] = files_list
        session_data['current_batch_id'] = batch_id
        save_user_session_data(session_data)
        
        if db is not None:
            flash(f'Files generated and stored in database! Batch ID: {batch_id}', 'success')
        else:
            flash('Files generated! (MongoDB not available - files stored in session only)', 'warning')
        
        return redirect(url_for('results'))
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('review'))

@app.route('/results')
def results():
    session_data = get_user_session_data()
    return render_template_string(RESULTS_TEMPLATE,
        output_files=session_data.get('output_files', []),
        audit_trail=session_data.get('audit_trail', []))

@app.route('/download/<filename>')
def download(filename):
    """Download file from MongoDB storage"""
    db = get_db()
    
    if db is not None:
        # Try to find file in MongoDB
        file_doc = db.output_files.find_one({'filename': filename})
        if file_doc:
            file_data = base64.b64decode(file_doc['file_data'])
            buffer = io.BytesIO(file_data)
            buffer.seek(0)
            return send_file(
                buffer,
                as_attachment=True,
                download_name=filename,
                mimetype=file_doc.get('content_type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            )
    
    flash('File not found in database. Please regenerate the files.', 'error')
    return redirect(url_for('results'))


@app.route('/download_all_zip')
def download_all_zip():
    """Download all output files as a single ZIP file"""
    import zipfile

    db = get_db()
    session_data = get_user_session_data()
    output_files = session_data.get('output_files', [])
    
    if not output_files:
        flash('No files to download. Please process a file first.', 'error')
        return redirect(url_for('results'))
    
    # Create ZIP file in memory
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for file_info in output_files:
            filename = file_info.get('filename')
            if filename and db is not None:
                # Get file from MongoDB
                file_doc = db.output_files.find_one({'filename': filename})
                if file_doc:
                    file_data = base64.b64decode(file_doc['file_data'])
                    zip_file.writestr(filename, file_data)
    
    zip_buffer.seek(0)
    
    # Generate ZIP filename with timestamp
    batch_id = session_data.get('current_batch_id', datetime.now().strftime('%Y%m%d_%H%M%S'))
    zip_filename = f"Bank_Transactions_{batch_id}.zip"
    
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=zip_filename,
        mimetype='application/zip'
    )


@app.route('/api/output-files')
def api_get_output_files():
    """Get list of all output files from MongoDB"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        # Get query parameters
        batch_id = request.args.get('batch_id')
        limit = int(request.args.get('limit', 50))
        
        query = {}
        if batch_id:
            query['batch_id'] = batch_id
        
        files = list(db.output_files.find(
            query, 
            {'file_data': 0}  # Exclude file data from listing
        ).sort('created_at', -1).limit(limit))
        
        return jsonify({
            'files': serialize_doc(files),
            'total': db.output_files.count_documents(query)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/output-files/<file_id>')
def api_get_output_file(file_id):
    """Download a specific output file by ID"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        from bson import ObjectId
        file_doc = db.output_files.find_one({'_id': ObjectId(file_id)})
        if not file_doc:
            return jsonify({'error': 'File not found'}), 404
        
        file_data = base64.b64decode(file_doc['file_data'])
        buffer = io.BytesIO(file_data)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=file_doc['filename'],
            mimetype=file_doc.get('content_type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/output-files/batch/<batch_id>')
def api_get_batch_files(batch_id):
    """Get all files for a specific batch"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        files = list(db.output_files.find(
            {'batch_id': batch_id},
            {'file_data': 0}
        ).sort('created_at', -1))
        
        return jsonify({
            'batch_id': batch_id,
            'files': serialize_doc(files),
            'count': len(files)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============ MONGODB REST API ENDPOINTS ============

@app.route('/api/status')
def api_status():
    """API health check and MongoDB status"""
    db = get_db()
    is_connected = db is not None
    return jsonify({
        'status': 'ok',
        'mongodb': 'connected' if is_connected else 'disconnected',
        'database': MONGODB_DATABASE if is_connected else None,
        'timestamp': datetime.now().isoformat()
    })

# ---------- TRANSACTIONS API ----------

@app.route('/api/transactions', methods=['GET'])
def api_get_transactions():
    """Get all transactions with optional filters"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        # Query parameters
        batch_id = request.args.get('batch_id')
        module = request.args.get('module')
        status = request.args.get('status')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        limit = int(request.args.get('limit', 100))
        skip = int(request.args.get('skip', 0))
        
        # Build query
        query = {}
        if batch_id:
            query['batch_id'] = batch_id
        if module:
            query['module'] = module.upper()
        if status:
            query['status'] = status
        if date_from or date_to:
            query['date'] = {}
            if date_from:
                query['date']['$gte'] = date_from
            if date_to:
                query['date']['$lte'] = date_to
        
        transactions = list(db.transactions.find(query).sort('date', -1).skip(skip).limit(limit))
        total = db.transactions.count_documents(query)
        
        return jsonify({
            'transactions': serialize_doc(transactions),
            'total': total,
            'limit': limit,
            'skip': skip
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/transactions', methods=['POST'])
def api_create_transaction():
    """Create a new transaction"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Validate required fields
        required = ['date', 'description', 'amount']
        for field in required:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Add metadata
        data['created_at'] = datetime.now()
        data['updated_at'] = datetime.now()
        data['status'] = data.get('status', 'pending')
        
        result = db.transactions.insert_one(data)
        
        # Log to audit
        db.audit_logs.insert_one({
            'action': 'create_transaction',
            'transaction_id': str(result.inserted_id),
            'timestamp': datetime.now(),
            'data': serialize_doc(data)
        })
        
        return jsonify({
            'status': 'success',
            'id': str(result.inserted_id),
            'message': 'Transaction created'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/transactions/<transaction_id>', methods=['GET'])
def api_get_transaction(transaction_id):
    """Get a single transaction by ID"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        transaction = db.transactions.find_one({'_id': ObjectId(transaction_id)})
        if transaction is None:
            return jsonify({'error': 'Transaction not found'}), 404
        return jsonify(serialize_doc(transaction))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/transactions/<transaction_id>', methods=['PUT'])
def api_update_transaction(transaction_id):
    """Update a transaction"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        data['updated_at'] = datetime.now()
        
        result = db.transactions.update_one(
            {'_id': ObjectId(transaction_id)},
            {'$set': data}
        )
        
        if result.matched_count == 0:
            return jsonify({'error': 'Transaction not found'}), 404
        
        # Log to audit
        db.audit_logs.insert_one({
            'action': 'update_transaction',
            'transaction_id': transaction_id,
            'timestamp': datetime.now(),
            'changes': serialize_doc(data)
        })
        
        return jsonify({'status': 'success', 'message': 'Transaction updated'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/transactions/<transaction_id>', methods=['DELETE'])
def api_delete_transaction(transaction_id):
    """Delete a transaction"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        result = db.transactions.delete_one({'_id': ObjectId(transaction_id)})
        if result.deleted_count == 0:
            return jsonify({'error': 'Transaction not found'}), 404
        
        # Log to audit
        db.audit_logs.insert_one({
            'action': 'delete_transaction',
            'transaction_id': transaction_id,
            'timestamp': datetime.now()
        })
        
        return jsonify({'status': 'success', 'message': 'Transaction deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------- BATCHES API ----------

@app.route('/api/batches', methods=['GET'])
def api_get_batches():
    """Get all transaction batches"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        limit = int(request.args.get('limit', 50))
        skip = int(request.args.get('skip', 0))
        
        batches = list(db.batches.find().sort('created_at', -1).skip(skip).limit(limit))
        total = db.batches.count_documents({})
        
        return jsonify({
            'batches': serialize_doc(batches),
            'total': total
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batches', methods=['POST'])
def api_create_batch():
    """Create a new batch with transactions"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        transactions = data.get('transactions', [])
        
        # Create batch record
        batch = {
            'name': data.get('name', f"Batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}"),
            'source_file': data.get('source_file', ''),
            'bank': data.get('bank', ''),
            'account': data.get('account', ''),
            'transaction_count': len(transactions),
            'total_credits': sum(t.get('amount', 0) for t in transactions if t.get('amount', 0) > 0),
            'total_debits': sum(abs(t.get('amount', 0)) for t in transactions if t.get('amount', 0) < 0),
            'status': 'pending',
            'created_at': datetime.now(),
            'updated_at': datetime.now()
        }
        
        batch_result = db.batches.insert_one(batch)
        batch_id = str(batch_result.inserted_id)
        
        # Insert transactions with batch_id
        if transactions:
            for txn in transactions:
                txn['batch_id'] = batch_id
                txn['created_at'] = datetime.now()
                txn['status'] = 'pending'
            db.transactions.insert_many(transactions)
        
        # Log to audit
        db.audit_logs.insert_one({
            'action': 'create_batch',
            'batch_id': batch_id,
            'transaction_count': len(transactions),
            'timestamp': datetime.now()
        })
        
        return jsonify({
            'status': 'success',
            'batch_id': batch_id,
            'transaction_count': len(transactions),
            'message': 'Batch created'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batches/<batch_id>', methods=['GET'])
def api_get_batch(batch_id):
    """Get a batch with its transactions"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        batch = db.batches.find_one({'_id': ObjectId(batch_id)})
        if batch is None:
            return jsonify({'error': 'Batch not found'}), 404
        
        transactions = list(db.transactions.find({'batch_id': batch_id}))
        
        result = serialize_doc(batch)
        result['transactions'] = serialize_doc(transactions)
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batches/<batch_id>/process', methods=['POST'])
def api_process_batch(batch_id):
    """Process a batch - classify and route transactions"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        batch = db.batches.find_one({'_id': ObjectId(batch_id)})
        if batch is None:
            return jsonify({'error': 'Batch not found'}), 404
        
        transactions = list(db.transactions.find({'batch_id': batch_id}))
        if not transactions:
            return jsonify({'error': 'No transactions in batch'}), 400
        
        # Classify transactions
        classifier = ClassificationEngine()
        classified = classifier.classify_batch([serialize_doc(t) for t in transactions])
        
        # Update transactions with classification
        for i, txn in enumerate(classified):
            db.transactions.update_one(
                {'_id': transactions[i]['_id']},
                {'$set': {
                    'module': txn.get('module', 'UNKNOWN'),
                    'category': txn.get('category', ''),
                    'gl_code': txn.get('gl_code', ''),
                    'fund_code': txn.get('fund_code', ''),
                    'confidence': txn.get('confidence', 0),
                    'status': 'classified',
                    'updated_at': datetime.now()
                }}
            )
        
        # Update batch status
        db.batches.update_one(
            {'_id': ObjectId(batch_id)},
            {'$set': {'status': 'classified', 'updated_at': datetime.now()}}
        )
        
        # Log to audit
        db.audit_logs.insert_one({
            'action': 'process_batch',
            'batch_id': batch_id,
            'transaction_count': len(transactions),
            'timestamp': datetime.now()
        })
        
        return jsonify({
            'status': 'success',
            'message': f'Processed {len(transactions)} transactions'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------- GL CODES API ----------

@app.route('/api/gl-codes', methods=['GET'])
def api_get_gl_codes():
    """Get all GL codes"""
    db = get_db()
    if db is None:
        # Return from static list if MongoDB not available
        return jsonify({'gl_codes': GL_CODES, 'source': 'static'})
    
    try:
        gl_codes = list(db.gl_codes.find({}, {'_id': 0}))
        if not gl_codes:
            # Initialize from static list
            if GL_CODES:
                db.gl_codes.insert_many(GL_CODES)
                gl_codes = GL_CODES
        return jsonify({'gl_codes': gl_codes, 'source': 'mongodb'})
    except Exception as e:
        return jsonify({'gl_codes': GL_CODES, 'source': 'static', 'error': str(e)})

@app.route('/api/gl-codes', methods=['POST'])
def api_create_gl_code():
    """Create a new GL code"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        data = request.get_json()
        if not data or 'code' not in data or 'name' not in data:
            return jsonify({'error': 'Code and name are required'}), 400
        
        # Check for duplicate
        if db.gl_codes.find_one({'code': data['code']}):
            return jsonify({'error': 'GL code already exists'}), 409
        
        data['created_at'] = datetime.now()
        db.gl_codes.insert_one(data)
        
        return jsonify({'status': 'success', 'message': 'GL code created'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/gl-codes/<code>', methods=['PUT'])
def api_update_gl_code(code):
    """Update a GL code"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        data = request.get_json()
        data['updated_at'] = datetime.now()
        
        result = db.gl_codes.update_one({'code': code}, {'$set': data})
        if result.matched_count == 0:
            return jsonify({'error': 'GL code not found'}), 404
        
        return jsonify({'status': 'success', 'message': 'GL code updated'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/gl-codes/<code>', methods=['DELETE'])
def api_delete_gl_code(code):
    """Delete a GL code"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        result = db.gl_codes.delete_one({'code': code})
        if result.deleted_count == 0:
            return jsonify({'error': 'GL code not found'}), 404
        
        return jsonify({'status': 'success', 'message': 'GL code deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------- FUND CODES API ----------

@app.route('/api/fund-codes', methods=['GET'])
def api_get_fund_codes():
    """Get all fund codes"""
    db = get_db()
    if db is None:
        return jsonify({'fund_codes': FUND_CODES, 'source': 'static'})
    
    try:
        fund_codes = list(db.fund_codes.find({}, {'_id': 0}))
        if not fund_codes:
            if FUND_CODES:
                db.fund_codes.insert_many(FUND_CODES)
                fund_codes = FUND_CODES
        return jsonify({'fund_codes': fund_codes, 'source': 'mongodb'})
    except Exception as e:
        return jsonify({'fund_codes': FUND_CODES, 'source': 'static', 'error': str(e)})

@app.route('/api/fund-codes', methods=['POST'])
def api_create_fund_code():
    """Create a new fund code"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        data = request.get_json()
        if not data or 'code' not in data or 'name' not in data:
            return jsonify({'error': 'Code and name are required'}), 400
        
        if db.fund_codes.find_one({'code': data['code']}):
            return jsonify({'error': 'Fund code already exists'}), 409
        
        data['created_at'] = datetime.now()
        db.fund_codes.insert_one(data)
        
        return jsonify({'status': 'success', 'message': 'Fund code created'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------- VENDORS API ----------

@app.route('/api/vendors', methods=['GET'])
def api_get_vendors():
    """Get all vendors"""
    db = get_db()
    if db is None:
        return jsonify({'vendors': get_all_vendors(), 'source': 'static'})
    
    try:
        vendors = list(db.vendors.find({}, {'_id': 0}))
        if not vendors:
            all_vendors = get_all_vendors()
            if all_vendors:
                db.vendors.insert_many(all_vendors)
                vendors = all_vendors
        return jsonify({'vendors': vendors, 'source': 'mongodb'})
    except Exception as e:
        return jsonify({'vendors': get_all_vendors(), 'source': 'static', 'error': str(e)})

@app.route('/api/vendors', methods=['POST'])
def api_create_vendor():
    """Create a new vendor"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        data = request.get_json()
        if not data or 'name' not in data:
            return jsonify({'error': 'Vendor name is required'}), 400
        
        if db.vendors.find_one({'name': data['name']}):
            return jsonify({'error': 'Vendor already exists'}), 409
        
        data['type'] = 'Vendor'
        data['created_at'] = datetime.now()
        db.vendors.insert_one(data)

        return jsonify({'status': 'success', 'message': 'Vendor created'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vendors/<name>', methods=['PUT'])
def api_update_vendor(name):
    """Update a vendor"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        data = request.get_json()
        data['updated_at'] = datetime.now()
        
        result = db.vendors.update_one({'name': name}, {'$set': data})
        if result.matched_count == 0:
            return jsonify({'error': 'Vendor not found'}), 404
        
        return jsonify({'status': 'success', 'message': 'Vendor updated'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vendors/<name>', methods=['DELETE'])
def api_delete_vendor(name):
    """Delete a vendor"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        result = db.vendors.delete_one({'name': name})
        if result.deleted_count == 0:
            return jsonify({'error': 'Vendor not found'}), 404
        
        return jsonify({'status': 'success', 'message': 'Vendor deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------- CUSTOMERS API ----------

@app.route('/api/customers', methods=['GET'])
def api_get_customers():
    """Get all customers"""
    db = get_db()
    if db is None:
        return jsonify({'customers': get_all_customers(), 'source': 'static'})
    
    try:
        customers = list(db.customers.find({}, {'_id': 0}))
        if not customers:
            all_customers = get_all_customers()
            if all_customers:
                db.customers.insert_many(all_customers)
                customers = all_customers
        return jsonify({'customers': customers, 'source': 'mongodb'})
    except Exception as e:
        return jsonify({'customers': get_all_customers(), 'source': 'static', 'error': str(e)})

@app.route('/api/customers', methods=['POST'])
def api_create_customer():
    """Create a new customer"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        data = request.get_json()
        if not data or 'name' not in data:
            return jsonify({'error': 'Customer name is required'}), 400
        
        if db.customers.find_one({'name': data['name']}):
            return jsonify({'error': 'Customer already exists'}), 409
        
        data['type'] = 'Customer'
        data['created_at'] = datetime.now()
        db.customers.insert_one(data)

        return jsonify({'status': 'success', 'message': 'Customer created'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/customers/<name>', methods=['PUT'])
def api_update_customer(name):
    """Update a customer"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        data = request.get_json()
        data['updated_at'] = datetime.now()
        
        result = db.customers.update_one({'name': name}, {'$set': data})
        if result.matched_count == 0:
            return jsonify({'error': 'Customer not found'}), 404
        
        return jsonify({'status': 'success', 'message': 'Customer updated'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/customers/<name>', methods=['DELETE'])
def api_delete_customer(name):
    """Delete a customer"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        result = db.customers.delete_one({'name': name})
        if result.deleted_count == 0:
            return jsonify({'error': 'Customer not found'}), 404
        
        return jsonify({'status': 'success', 'message': 'Customer deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------- AUDIT LOGS API ----------

@app.route('/api/audit-logs', methods=['GET'])
def api_get_audit_logs():
    """Get audit logs"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        limit = int(request.args.get('limit', 100))
        skip = int(request.args.get('skip', 0))
        action = request.args.get('action')
        
        query = {}
        if action:
            query['action'] = action
        
        logs = list(db.audit_logs.find(query).sort('timestamp', -1).skip(skip).limit(limit))
        total = db.audit_logs.count_documents(query)
        
        return jsonify({
            'audit_logs': serialize_doc(logs),
            'total': total
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------- SYNC API ----------

@app.route('/api/sync/master-data', methods=['POST'])
def api_sync_master_data():
    """Sync static master data (GL codes, Fund codes, Vendors, Customers) to MongoDB"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        results = {}
        
        # Sync GL Codes
        db.gl_codes.delete_many({})
        if GL_CODES:
            db.gl_codes.insert_many(GL_CODES)
        results['gl_codes'] = len(GL_CODES)
        
        # Sync Fund Codes
        db.fund_codes.delete_many({})
        if FUND_CODES:
            db.fund_codes.insert_many(FUND_CODES)
        results['fund_codes'] = len(FUND_CODES)
        
        # Sync Vendors
        all_vendors = get_all_vendors()
        db.vendors.delete_many({})
        if all_vendors:
            db.vendors.insert_many(all_vendors)
        results['vendors'] = len(all_vendors)
        
        # Sync Customers
        all_customers = get_all_customers()
        db.customers.delete_many({})
        if all_customers:
            db.customers.insert_many(all_customers)
        results['customers'] = len(all_customers)
        
        # Log to audit
        db.audit_logs.insert_one({
            'action': 'sync_master_data',
            'timestamp': datetime.now(),
            'results': results
        })
        
        return jsonify({
            'status': 'success',
            'message': 'Master data synced to MongoDB',
            'counts': results
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------- STATISTICS API ----------

@app.route('/api/stats', methods=['GET'])
def api_get_stats():
    """Get dashboard statistics"""
    db = get_db()
    if db is None:
        return jsonify({'error': 'MongoDB not available'}), 503
    
    try:
        stats = {
            'transactions': {
                'total': db.transactions.count_documents({}),
                'pending': db.transactions.count_documents({'status': 'pending'}),
                'classified': db.transactions.count_documents({'status': 'classified'}),
                'processed': db.transactions.count_documents({'status': 'processed'})
            },
            'batches': {
                'total': db.batches.count_documents({}),
                'pending': db.batches.count_documents({'status': 'pending'}),
                'classified': db.batches.count_documents({'status': 'classified'}),
                'processed': db.batches.count_documents({'status': 'processed'})
            },
            'master_data': {
                'gl_codes': db.gl_codes.count_documents({}),
                'fund_codes': db.fund_codes.count_documents({}),
                'vendors': db.vendors.count_documents({}),
                'customers': db.customers.count_documents({})
            },
            'by_module': {
                'CR': db.transactions.count_documents({'module': 'CR'}),
                'CD': db.transactions.count_documents({'module': 'CD'}),
                'JV': db.transactions.count_documents({'module': 'JV'}),
                'UNKNOWN': db.transactions.count_documents({'module': 'UNKNOWN'})
            }
        }
        
        # Calculate totals
        pipeline = [
            {'$group': {
                '_id': None,
                'total_credits': {'$sum': {'$cond': [{'$gt': ['$amount', 0]}, '$amount', 0]}},
                'total_debits': {'$sum': {'$cond': [{'$lt': ['$amount', 0]}, {'$abs': '$amount'}, 0]}}
            }}
        ]
        totals = list(db.transactions.aggregate(pipeline))
        if totals:
            stats['totals'] = {
                'credits': totals[0].get('total_credits', 0),
                'debits': totals[0].get('total_debits', 0)
            }
        
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============ CHROMADB LEARNING API ============

# Import learning module (with fallback if not available)
# ChromaDB may be installed in C:\py_libs due to Windows long path issues
try:
    # Try standard import first
    from learning import get_gl_suggester, get_chroma_store
    LEARNING_AVAILABLE = True
    print("[INFO] ChromaDB Learning module loaded successfully.")
except ImportError:
    # Try with custom library path
    try:
        import sys
        if 'C:\\py_libs' not in sys.path:
            sys.path.insert(0, 'C:\\py_libs')
        from learning import get_gl_suggester, get_chroma_store
        LEARNING_AVAILABLE = True
        print("[INFO] ChromaDB Learning module loaded from C:\\py_libs.")
    except ImportError as e:
        LEARNING_AVAILABLE = False
        print(f"[INFO] Learning module not available: {e}. ChromaDB learning features disabled.")

@app.route('/api/learning/suggest', methods=['POST'])
def api_suggest_gl_code():
    """
    Get GL code suggestion for a transaction description.

    Request body:
    {
        "description": "HUD TREAS NAHASDA",
        "type": "deposit" or "withdrawal",
        "bank": "PNC" (optional)
    }

    Response:
    {
        "gl_code": "3001",
        "gl_name": "Revenue - Federal",
        "confidence": 92.5,
        "confidence_level": "high",
        "source": "learned" | "keyword" | "default",
        "reason": "Similar to: HUD TREASURY PAYMENT..."
    }
    """
    if not LEARNING_AVAILABLE:
        return jsonify({'error': 'Learning module not available'}), 503

    try:
        data = request.get_json() or {}
        description = data.get('description', '')
        transaction_type = data.get('type', 'withdrawal')
        bank = data.get('bank')
        amount = data.get('amount')

        if not description:
            return jsonify({'error': 'Description is required'}), 400

        suggester = get_gl_suggester()
        result = suggester.suggest(
            description=description,
            transaction_type=transaction_type,
            amount=amount,
            bank=bank
        )

        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/learning/learn', methods=['POST'])
def api_learn_transaction():
    """
    Learn from a user-approved transaction.

    Request body:
    {
        "description": "NEW VENDOR PAYMENT",
        "gl_code": "7300",
        "type": "withdrawal",
        "module": "CD",
        "bank": "PNC",
        "user_corrected": false
    }
    """
    if not LEARNING_AVAILABLE:
        return jsonify({'error': 'Learning module not available'}), 503

    try:
        data = request.get_json() or {}

        required = ['description', 'gl_code', 'type', 'module']
        for field in required:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400

        suggester = get_gl_suggester()
        doc_id = suggester.learn_from_user(
            description=data['description'],
            gl_code=data['gl_code'],
            transaction_type=data['type'],
            module=data['module'],
            bank=data.get('bank', 'Unknown'),
            user_corrected=data.get('user_corrected', False)
        )

        return jsonify({
            'success': True,
            'doc_id': doc_id,
            'message': f"Learned pattern for GL {data['gl_code']}"
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/learning/learn-batch', methods=['POST'])
def api_learn_batch():
    """
    Learn from multiple approved transactions at once.

    Request body:
    {
        "transactions": [
            {"description": "...", "gl_code": "...", "type": "...", "module": "..."},
            ...
        ],
        "bank": "PNC"
    }
    """
    if not LEARNING_AVAILABLE:
        return jsonify({'error': 'Learning module not available'}), 503

    try:
        data = request.get_json() or {}
        transactions = data.get('transactions', [])
        bank = data.get('bank', 'Unknown')

        if not transactions:
            return jsonify({'error': 'No transactions provided'}), 400

        # Add bank to each transaction
        for txn in transactions:
            txn['bank'] = bank

        suggester = get_gl_suggester()
        count = suggester.learn_batch(transactions, bank)

        return jsonify({
            'success': True,
            'learned_count': count,
            'message': f"Learned from {count} transactions"
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/learning/stats', methods=['GET'])
def api_learning_stats():
    """Get learning statistics."""
    if not LEARNING_AVAILABLE:
        return jsonify({'error': 'Learning module not available'}), 503

    try:
        suggester = get_gl_suggester()
        stats = suggester.get_learning_stats()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/learning/export', methods=['GET'])
def api_export_patterns():
    """Export all learned patterns to JSON file."""
    if not LEARNING_AVAILABLE:
        return jsonify({'error': 'Learning module not available'}), 503

    try:
        export_dir = os.path.join(DATA_DIR, 'exports')
        os.makedirs(export_dir, exist_ok=True)

        filepath = os.path.join(export_dir, 'learned_patterns.json')

        store = get_chroma_store()
        count = store.export_patterns(filepath)

        return jsonify({
            'success': True,
            'count': count,
            'filepath': filepath
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/learning/import', methods=['POST'])
def api_import_patterns():
    """Import patterns from JSON file."""
    if not LEARNING_AVAILABLE:
        return jsonify({'error': 'Learning module not available'}), 503

    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if not file.filename.endswith('.json'):
            return jsonify({'error': 'Only JSON files supported'}), 400

        # Save to temp file (in-memory style)
        import tempfile
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.json', delete=False) as tmp:
            file.save(tmp.name)
            temp_path = tmp.name

        try:
            store = get_chroma_store()
            count = store.import_patterns(temp_path)
        finally:
            # Clean up temp file
            try:
                os.unlink(temp_path)
            except:
                pass

        return jsonify({
            'success': True,
            'imported_count': count,
            'message': f"Imported {count} patterns"
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/learning/keywords', methods=['GET'])
def api_get_keywords():
    """Get current keyword rules."""
    if not LEARNING_AVAILABLE:
        return jsonify({'error': 'Learning module not available'}), 503

    try:
        suggester = get_gl_suggester()
        return jsonify(suggester.keywords)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/learning/keywords', methods=['POST'])
def api_add_keyword():
    """
    Add a new keyword rule.

    Request body:
    {
        "keyword": "NEW COMPANY",
        "gl_code": "7300",
        "gl_name": "Accounts Payable",
        "type": "withdrawal"
    }
    """
    if not LEARNING_AVAILABLE:
        return jsonify({'error': 'Learning module not available'}), 503

    try:
        data = request.get_json() or {}

        required = ['keyword', 'gl_code', 'gl_name', 'type']
        for field in required:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400

        suggester = get_gl_suggester()
        success = suggester.add_keyword_rule(
            keyword=data['keyword'],
            gl_code=data['gl_code'],
            gl_name=data['gl_name'],
            transaction_type=data['type']
        )

        if success:
            return jsonify({
                'success': True,
                'message': f"Added keyword rule: {data['keyword']} -> GL {data['gl_code']}"
            })
        else:
            return jsonify({'error': 'Failed to save keyword'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============ TEMPLATES ============

INDEX_TEMPLATE = '''<!DOCTYPE html>
<html><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Bank Transaction Posting Tool</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.0/font/bootstrap-icons.css" rel="stylesheet">
<style>.drag-drop{border:2px dashed #ccc;border-radius:10px;padding:40px;text-align:center}.drag-drop.dragover{border-color:#0d6efd;background:#e7f1ff}</style>
</head><body>
<nav class="navbar navbar-dark bg-primary"><div class="container"><a class="navbar-brand" href="/"><i class="bi bi-bank"></i> Bank Transaction Posting Tool</a></div></nav>
<div class="container mt-4">
{% with messages = get_flashed_messages(with_categories=true) %}{% if messages %}{% for cat, msg in messages %}<div class="alert alert-{{ 'danger' if cat == 'error' else cat }} alert-dismissible fade show">{{ msg }}<button type="button" class="btn-close" data-bs-dismiss="alert"></button></div>{% endfor %}{% endif %}{% endwith %}
<div class="row justify-content-center"><div class="col-md-8">
<div class="card"><div class="card-header bg-primary text-white"><h4><i class="bi bi-upload"></i> Upload Bank Statement</h4></div>
<div class="card-body">
<form action="/upload" method="post" enctype="multipart/form-data">
<div class="drag-drop" id="dropZone" onclick="document.getElementById('fileInput').click()">
<i class="bi bi-cloud-upload" style="font-size:3em;color:#6c757d"></i>
<h5 class="mt-3">Drag & Drop or Click to Browse</h5>
<p class="text-muted">PDF, Excel, CSV supported</p>
<input type="file" name="file" id="fileInput" class="d-none" accept=".pdf,.xlsx,.xls,.csv">
</div>
<div id="selectedFile" class="mt-3 d-none"><div class="alert alert-info"><i class="bi bi-file-earmark"></i> <span id="fileName"></span>
<button type="submit" class="btn btn-success btn-sm float-end"><i class="bi bi-play-fill"></i> Process</button></div></div>
</form></div></div></div></div></div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
<script>
const dz=document.getElementById('dropZone'),fi=document.getElementById('fileInput');
dz.ondragover=e=>{e.preventDefault();dz.classList.add('dragover')};
dz.ondragleave=()=>dz.classList.remove('dragover');
dz.ondrop=e=>{e.preventDefault();dz.classList.remove('dragover');if(e.dataTransfer.files.length){fi.files=e.dataTransfer.files;show(e.dataTransfer.files[0].name)}};
fi.onchange=()=>{if(fi.files.length)show(fi.files[0].name)};
function show(n){document.getElementById('fileName').textContent=n;document.getElementById('selectedFile').classList.remove('d-none')}
</script></body></html>'''

REVIEW_TEMPLATE = '''<!DOCTYPE html>
<html><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Review - Bank Transaction Posting Tool</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.0/font/bootstrap-icons.css" rel="stylesheet">
<style>
.badge-high{background:#28a745}.badge-medium{background:#ffc107;color:#000}.badge-low,.badge-none{background:#dc3545}.badge-manual{background:#0d6efd}
.needs-review{background:#fff3cd}.selected{background:#cfe2ff!important}
.suggestion-hint{background:#e7f3ff;border-left:3px solid #0d6efd;padding:10px;margin-bottom:15px}
.vendor-match{background:#f8f9fa;border:1px solid #dee2e6;border-radius:8px;padding:15px;margin-top:10px}
.vendor-item{padding:10px;margin:5px 0;background:#fff;border:1px solid #ddd;border-radius:4px;cursor:pointer}.vendor-item:hover{border-color:#0d6efd;background:#e7f1ff}
.audit-item{font-size:.85em;padding:5px 10px;border-left:3px solid #6c757d;margin-bottom:5px;background:#f8f9fa}
/* Select2 customization */
.select2-container{width:100%!important}
.select2-container .select2-selection--single{height:38px;border:1px solid #ced4da;border-radius:0.375rem}
.select2-container--default .select2-selection--single .select2-selection__rendered{line-height:36px;padding-left:12px}
.select2-container--default .select2-selection--single .select2-selection__arrow{height:36px}
.select2-dropdown{border:1px solid #ced4da;border-radius:0.375rem}
.select2-search--dropdown .select2-search__field{border:1px solid #ced4da;border-radius:0.375rem;padding:6px 12px}
.select2-results__option--highlighted[aria-selected]{background-color:#0d6efd!important}
</style>
<!-- jQuery and Select2 for searchable dropdowns -->
<script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
<link href="https://cdn.jsdelivr.net/npm/select2@4.1.0-rc.0/dist/css/select2.min.css" rel="stylesheet" />
<script src="https://cdn.jsdelivr.net/npm/select2@4.1.0-rc.0/dist/js/select2.min.js"></script>
</head><body>
<nav class="navbar navbar-dark bg-primary"><div class="container"><a class="navbar-brand" href="/"><i class="bi bi-bank"></i> Bank Transaction Posting Tool</a>
<div class="navbar-nav ms-auto"><a class="nav-link" href="/">Upload</a><a class="nav-link active" href="/review">Review</a><a class="nav-link" href="/results">Results</a></div></div></nav>
<div class="container mt-4">
{% with messages = get_flashed_messages(with_categories=true) %}{% if messages %}{% for cat,msg in messages %}<div class="alert alert-{{ 'danger' if cat=='error' else cat }} alert-dismissible fade show">{{ msg }}<button type="button" class="btn-close" data-bs-dismiss="alert"></button></div>{% endfor %}{% endif %}{% endwith %}

<!-- Validation Warnings Section -->
{% if parsing_metadata and parsing_metadata.warnings %}
<div class="alert alert-warning alert-dismissible fade show" role="alert">
<h5 class="alert-heading"><i class="bi bi-exclamation-triangle"></i> Validation Warnings</h5>
<ul class="mb-0">
{% for warning in parsing_metadata.warnings %}
<li><span class="badge bg-{{ 'danger' if warning.severity == 'high' else 'warning' }}">{{ warning.severity|upper }}</span> {{ warning.message }}</li>
{% endfor %}
</ul>
{% if parsing_metadata.ocr_used %}<hr><small class="text-muted"><i class="bi bi-info-circle"></i> OCR was used to extract text from this document. Some values may need verification.</small>{% endif %}
<button type="button" class="btn-close" data-bs-dismiss="alert"></button>
</div>
{% endif %}

<!-- Bank Summary Comparison (if available) -->
{% if parsing_metadata and parsing_metadata.expected_deposits %}
<div class="card mb-3 border-info">
<div class="card-header bg-info text-white"><i class="bi bi-file-earmark-check"></i> Bank Statement Validation</div>
<div class="card-body">
<div class="row">
<div class="col-md-6">
<table class="table table-sm table-bordered mb-0">
<thead class="table-light"><tr><th></th><th>Expected (Bank)</th><th>Parsed</th><th>Match</th></tr></thead>
<tbody>
<tr><td><strong>Deposits</strong></td>
<td>${{ '{:,.2f}'.format(parsing_metadata.expected_deposits) }}</td>
<td>${{ '{:,.2f}'.format(parsing_metadata.parsed_deposits) }}</td>
<td>{% if (parsing_metadata.parsed_deposits - parsing_metadata.expected_deposits)|abs < 1 %}<span class="badge bg-success">✓</span>{% else %}<span class="badge bg-danger">✗</span>{% endif %}</td>
</tr>
{% if parsing_metadata.expected_withdrawals %}
<tr><td><strong>Withdrawals</strong></td>
<td>${{ '{:,.2f}'.format(parsing_metadata.expected_withdrawals) }}</td>
<td>${{ '{:,.2f}'.format(parsing_metadata.parsed_withdrawals) }}</td>
<td>{% if (parsing_metadata.parsed_withdrawals - parsing_metadata.expected_withdrawals)|abs < 1 %}<span class="badge bg-success">✓</span>{% else %}<span class="badge bg-danger">✗</span>{% endif %}</td>
</tr>
{% endif %}
</tbody>
</table>
</div>
<div class="col-md-6">
<div class="p-3 bg-light rounded">
<p class="mb-1"><strong>Bank:</strong> {{ parsing_metadata.bank_name or 'Unknown' }}</p>
<p class="mb-1"><strong>Statement Period:</strong> {{ parsing_metadata.statement_period or (parsing_metadata.statement_month|string + '/' + parsing_metadata.statement_year|string if parsing_metadata.statement_month else 'N/A') }}</p>
<p class="mb-1"><strong>OCR Used:</strong> {% if parsing_metadata.ocr_used %}<span class="badge bg-warning text-dark">Yes</span>{% else %}<span class="badge bg-secondary">No</span>{% endif %}</p>
{% if parsing_metadata.quality_score is defined %}
<p class="mb-0"><strong>Quality Score:</strong>
{% if parsing_metadata.quality_score >= 80 %}<span class="badge bg-success">{{ parsing_metadata.quality_score }}/100</span>
{% elif parsing_metadata.quality_score >= 60 %}<span class="badge bg-warning text-dark">{{ parsing_metadata.quality_score }}/100</span>
{% else %}<span class="badge bg-danger">{{ parsing_metadata.quality_score }}/100</span>{% endif %}
</p>
{% endif %}
</div>
</div>
</div>
</div>
</div>
{% endif %}

<div class="row mb-4">
<div class="col-md-3"><div class="card text-center bg-success text-white p-3 kpi-card" style="cursor:pointer" onclick="filterByModule('cr')" title="Click to view Cash Receipts"><h2>{{ by_module.CR|length }}</h2><div>Cash Receipts</div><small class="opacity-75">${{ '{:,.0f}'.format(summary.total_credits) }}</small></div></div>
<div class="col-md-3"><div class="card text-center bg-danger text-white p-3 kpi-card" style="cursor:pointer" onclick="filterByModule('cd')" title="Click to view Cash Disbursements"><h2>{{ by_module.CD|length }}</h2><div>Cash Disbursements</div><small class="opacity-75">${{ '{:,.0f}'.format(summary.total_debits) }}</small></div></div>
<div class="col-md-3"><div class="card text-center bg-warning p-3 kpi-card" style="cursor:pointer" onclick="filterByModule('jv')" title="Click to view Journal Vouchers"><h2>{{ by_module.JV|length }}</h2><div>Journal Vouchers</div><small class="opacity-75">${{ '{:,.0f}'.format(summary.total_jv) if summary.total_jv else 0 }}</small></div></div>
<div class="col-md-3"><div class="card text-center bg-secondary text-white p-3 kpi-card" style="cursor:pointer" onclick="filterByModule('unk')" title="Click to view Unidentified"><h2>{{ by_module.UNKNOWN|length }}</h2><div>Unidentified</div><small class="opacity-75">{% if by_module.UNKNOWN|length > 0 %}Needs Review{% else %}All Good{% endif %}</small></div></div>
</div>

<!-- Summary Section - Moved to Top -->
<div class="card mb-4"><div class="card-header bg-dark text-white"><h5 class="mb-0"><i class="bi bi-calculator"></i> Summary</h5></div>
<div class="card-body"><div class="row">
<div class="col-md-4"><div class="card border-success"><div class="card-body text-center"><h6 class="text-muted">Total Deposits</h6><h3 class="text-success">${{ '{:,.2f}'.format(summary.total_credits) }}</h3></div></div></div>
<div class="col-md-4"><div class="card border-danger"><div class="card-body text-center"><h6 class="text-muted">Total Withdrawals</h6><h3 class="text-danger">${{ '{:,.2f}'.format(summary.total_debits) }}</h3></div></div></div>
<div class="col-md-4"><div class="card border-info"><div class="card-body text-center"><h6 class="text-muted">Net Cash Flow</h6><h3 class="{{ 'text-success' if summary.balance >= 0 else 'text-danger' }}">{% if summary.balance >= 0 %}↑${{ '{:,.2f}'.format(summary.balance) }}{% else %}↓${{ '{:,.2f}'.format(summary.balance|abs) }}{% endif %}</h3></div></div></div>
</div>
{% if by_year and by_year|length > 1 %}
<hr class="my-3">
<h6 class="text-muted mb-2"><i class="bi bi-calendar3"></i> By Fiscal Year</h6>
<div class="row">
{% for year, data in by_year|dictsort(reverse=true) %}
<div class="col-md-{{ 12 // (by_year|length) if by_year|length <= 4 else 3 }}">
<div class="card bg-light mb-2"><div class="card-body py-2 px-3">
<h6 class="mb-1 fw-bold">FY {{ year }}</h6>
<small class="d-block text-success">Deposits: ${{ '{:,.2f}'.format(data.deposits) }}</small>
<small class="d-block text-danger">Withdrawals: ${{ '{:,.2f}'.format(data.withdrawals) }}</small>
<small class="d-block text-muted">{{ data.count }} transaction{{ 's' if data.count != 1 else '' }}</small>
</div></div>
</div>
{% endfor %}
</div>
{% endif %}
</div></div>

<div class="card mb-3 d-none" id="bulkBar"><div class="card-body bg-light">
<div class="row align-items-center">
<div class="col-auto"><strong><span id="selCount">0</span> selected</strong></div>
<div class="col-auto"><select class="form-select form-select-sm" id="bulkModule" style="width:auto"><option value="">Module...</option><option value="CR">CR</option><option value="CD">CD</option><option value="JV">JV</option></select></div>
<div class="col-auto"><select class="form-select form-select-sm" id="bulkGL" style="width:auto"><option value="">GL Code...</option>{% for g in gl_codes %}<option value="{{ g.code }}">{{ g.code }}-{{ g.name[:20] }}</option>{% endfor %}</select></div>
<div class="col-auto"><select class="form-select form-select-sm" id="bulkFund" style="width:auto"><option value="">Fund...</option>{% for f in fund_codes %}<option value="{{ f.code }}">{{ f.name }}</option>{% endfor %}</select></div>
<div class="col-auto"><button class="btn btn-primary btn-sm" onclick="applyBulk()"><i class="bi bi-check-all"></i> Apply</button>
<button class="btn btn-outline-secondary btn-sm" onclick="clearSel()"><i class="bi bi-x"></i> Clear</button></div>
</div></div></div>

<div class="card"><div class="card-header d-flex justify-content-between align-items-center">
<h5><i class="bi bi-list-check"></i> Review Transactions</h5>
<div>{% if audit_trail %}<button class="btn btn-outline-secondary btn-sm me-2" onclick="undoLast()"><i class="bi bi-arrow-counterclockwise"></i> Undo</button>{% endif %}
<button class="btn btn-outline-primary btn-sm me-2" data-bs-toggle="modal" data-bs-target="#addTransactionModal"><i class="bi bi-plus-circle"></i> Add Transaction</button>
<button class="btn btn-outline-success btn-sm me-2" data-bs-toggle="modal" data-bs-target="#addCustomerModal"><i class="bi bi-person-plus"></i> Add Customer</button>
<button class="btn btn-outline-info btn-sm me-2" data-bs-toggle="modal" data-bs-target="#addVendorModal"><i class="bi bi-building-add"></i> Add Vendor</button>
<form action="/process" method="post" class="d-inline"><button type="submit" class="btn btn-primary"><i class="bi bi-gear"></i> Generate Files</button></form></div>
</div>
<div class="card-body">
<ul class="nav nav-tabs" role="tablist">
<li class="nav-item"><button class="nav-link active" data-bs-toggle="tab" data-bs-target="#all">All ({{ transactions|length }})</button></li>
<li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#cr">CR ({{ by_module.CR|length }})</button></li>
<li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#cd">CD ({{ by_module.CD|length }})</button></li>
<li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#jv">JV ({{ by_module.JV|length }})</button></li>
<li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#unk">Unknown ({{ by_module.UNKNOWN|length }})</button></li>
</ul>
<div class="tab-content mt-3">
<div class="tab-pane fade show active" id="all">
<table class="table table-hover table-sm">
<thead class="table-light"><tr><th style="width:40px"><input type="checkbox" id="selAll" onchange="toggleAll()"></th><th>Date</th><th>Description</th><th>Amount</th><th>Module</th><th>Customer/Vendor</th><th>GL</th><th>Fund</th><th>Confidence</th><th>Actions</th></tr></thead>
<tbody>{% for t in transactions %}
<tr class="{{ 'needs-review' if t.confidence_level in ['low','none'] }}" data-idx="{{ loop.index0 }}">
<td><input type="checkbox" class="rowCb" value="{{ loop.index0 }}" onchange="updateSel()"></td>
<td>{{ t.date }}</td>
<td title="{{ t.description }}">{{ t.description[:30] }}{% if t.description|length > 30 %}...{% endif %}</td>
<td class="{{ 'text-success' if t.amount > 0 else 'text-danger' }}">${{ '{:,.2f}'.format(t.amount|abs) }}</td>
<td><span class="badge bg-{{ 'success' if t.module=='CR' else 'danger' if t.module=='CD' else 'warning' if t.module=='JV' else 'secondary' }}">{{ t.module }}</span></td>
<td><span class="text-primary">{{ t.payee or t.vendor or '-' }}</span></td>
<td>{{ t.gl_code or '-' }}</td>
<td>{{ t.fund_code or '-' }}</td>
<td><span class="badge badge-{{ t.confidence_level }}">{{ t.confidence_level }}</span></td>
<td><button class="btn btn-sm btn-outline-primary edit-btn" data-idx="{{ loop.index0 }}" data-desc="{{ t.description }}" data-module="{{ t.module }}" data-gl="{{ t.gl_code or '' }}" data-fund="{{ t.fund_code or '' }}" data-payee="{{ t.payee or t.vendor or '' }}" data-amt="{{ t.amount }}"><i class="bi bi-pencil"></i></button>
<button class="btn btn-sm btn-outline-danger delete-btn ms-1" data-idx="{{ loop.index0 }}" data-desc="{{ t.description }}" title="Delete"><i class="bi bi-trash"></i></button></td>
</tr>{% endfor %}</tbody>
</table></div>
<div class="tab-pane fade" id="cr">{% if by_module.CR %}<table class="table table-sm"><thead><tr><th>Date</th><th>Description</th><th>Amount</th><th>Customer/Vendor</th><th>GL</th><th>Fund</th></tr></thead><tbody>{% for t in by_module.CR %}<tr><td>{{ t.date }}</td><td>{{ t.description[:35] }}</td><td class="text-success">${{ '{:,.2f}'.format(t.amount) }}</td><td class="text-primary">{{ t.payee or t.vendor or '-' }}</td><td>{{ t.gl_code or '-' }}</td><td>{{ t.fund_code or '-' }}</td></tr>{% endfor %}</tbody></table>{% else %}<div class="alert alert-info">No CR transactions</div>{% endif %}</div>
<div class="tab-pane fade" id="cd">{% if by_module.CD %}<table class="table table-sm"><thead><tr><th>Date</th><th>Description</th><th>Amount</th><th>Customer/Vendor</th><th>GL</th><th>Fund</th></tr></thead><tbody>{% for t in by_module.CD %}<tr><td>{{ t.date }}</td><td>{{ t.description[:35] }}</td><td class="text-danger">${{ '{:,.2f}'.format(t.amount|abs) }}</td><td class="text-primary">{{ t.payee or t.vendor or '-' }}</td><td>{{ t.gl_code or '-' }}</td><td>{{ t.fund_code or '-' }}</td></tr>{% endfor %}</tbody></table>{% else %}<div class="alert alert-info">No CD transactions</div>{% endif %}</div>
<div class="tab-pane fade" id="jv">{% if by_module.JV %}<table class="table table-sm"><thead><tr><th>Date</th><th>Description</th><th>Amount</th><th>Customer/Vendor</th><th>GL</th><th>Fund</th></tr></thead><tbody>{% for t in by_module.JV %}<tr><td>{{ t.date }}</td><td>{{ t.description[:35] }}</td><td>${{ '{:,.2f}'.format(t.amount|abs) }}</td><td class="text-primary">{{ t.payee or t.vendor or '-' }}</td><td>{{ t.gl_code or '-' }}</td><td>{{ t.fund_code or '-' }}</td></tr>{% endfor %}</tbody></table>{% else %}<div class="alert alert-info">No JV transactions</div>{% endif %}</div>
<div class="tab-pane fade" id="unk">{% if by_module.UNKNOWN %}<table class="table table-sm"><thead><tr><th>Date</th><th>Description</th><th>Amount</th></tr></thead><tbody>{% for t in by_module.UNKNOWN %}<tr class="needs-review"><td>{{ t.date }}</td><td>{{ t.description[:35] }}</td><td>${{ '{:,.2f}'.format(t.amount|abs) }}</td></tr>{% endfor %}</tbody></table>{% else %}<div class="alert alert-success"><i class="bi bi-check-circle"></i> All classified!</div>{% endif %}</div>
</div></div></div>

{% if audit_trail %}<div class="card mt-4 border-info">
<div class="card-header bg-info text-white d-flex justify-content-between align-items-center">
<h5 class="mb-0"><i class="bi bi-clock-history"></i> Audit Trail ({{ audit_trail|length }} changes)</h5>
<small>Session: {{ session_id }}</small>
</div>
<div class="card-body p-0" style="max-height:200px;overflow-y:auto">
<table class="table table-sm table-hover mb-0">
<thead class="table-light sticky-top"><tr><th style="width:120px">Timestamp</th><th>Transaction</th><th>Changes</th><th style="width:80px">Action</th></tr></thead>
<tbody>{% for e in audit_trail|reverse %}
<tr><td><small class="text-muted">{{ e.timestamp }}</small></td>
<td><strong>{{ e.transaction[:40] }}</strong></td>
<td><small>{{ e.changes|join(' | ') }}</small></td>
<td><span class="badge bg-{{ 'success' if 'Module' in (e.changes|join('')) else ('warning' if 'GL' in (e.changes|join('')) else 'secondary') }}">{{ e.action if e.action else 'Edit' }}</span></td>
</tr>{% endfor %}
</tbody></table>
</div>
<div class="card-footer bg-light text-muted small"><i class="bi bi-info-circle"></i> All changes are tracked for compliance. Click Undo to revert last change.</div>
</div>{% endif %}

<!-- Edit Modal -->
<div class="modal fade" id="editModal" tabindex="-1"><div class="modal-dialog modal-lg"><div class="modal-content">
<div class="modal-header"><h5><i class="bi bi-pencil-square"></i> Edit Transaction</h5><button type="button" class="btn-close" data-bs-dismiss="modal"></button></div>
<div class="modal-body">
<div class="suggestion-hint" id="suggHint"><strong><i class="bi bi-lightbulb"></i> AI Suggestion:</strong> <span id="suggText">-</span>
<button class="btn btn-sm btn-outline-primary float-end" onclick="applySugg()"><i class="bi bi-check"></i> Apply</button></div>
<div class="mb-3 p-2 bg-light rounded"><strong>Description:</strong> <span id="editDesc"></span> | <strong>Amount:</strong> <span id="editAmt"></span></div>
<form id="editForm"><input type="hidden" id="editIdx">
<div class="row">
<div class="col-md-6 mb-3">
<label class="form-label">Module</label>
<select class="form-select" id="editModule"><option value="CR">Cash Receipt (CR)</option><option value="CD">Cash Disbursement (CD)</option><option value="JV">Journal Voucher (JV)</option></select>
<small class="text-muted" id="modExplain"></small>
</div>
<div class="col-md-6 mb-3">
<label class="form-label">Customer/Vendor <small class="text-muted">(auto-fills GL & Fund)</small></label>
<select class="form-select" id="editPayeeSelect" onchange="onPayeeSelect()">
<option value="">-- Select or type below --</option>
<optgroup label="Customers">{% for c in customers %}<option value="{{ c.name }}" data-gl="{{ c.gl_code }}" data-fund="{{ c.fund }}">{{ c.name }}</option>{% endfor %}</optgroup>
<optgroup label="Vendors">{% for v in vendors %}<option value="{{ v.name }}" data-gl="{{ v.gl_code }}" data-fund="{{ v.fund }}">{{ v.name }}</option>{% endfor %}</optgroup>
</select>
</div>
</div>
<div class="row">
<div class="col-md-4 mb-3">
<label class="form-label">GL Code</label>
<select class="form-select" id="editGL">
<option value="">-- Select --</option>
{% for g in gl_codes %}<option value="{{ g.code }}">{{ g.code }} - {{ g.name[:25] }}</option>{% endfor %}
</select>
</div>
<div class="col-md-4 mb-3">
<label class="form-label">Fund/Class</label>
<select class="form-select" id="editFund">
<option value="">-- Select --</option>
{% for f in fund_codes %}<option value="{{ f.code }}">{{ f.name }}</option>{% endfor %}
</select>
</div>
<div class="col-md-4 mb-3">
<label class="form-label">Payee Name</label>
<input type="text" class="form-control" id="editPayee" placeholder="Or type custom">
</div>
</div>
</form>
<div class="vendor-match d-none" id="vendorPanel"><h6><i class="bi bi-search"></i> Matches from Description:</h6><div id="vendorMatches"></div></div>
</div>
<div class="modal-footer"><button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Cancel</button><button type="button" class="btn btn-primary" id="saveEdit"><i class="bi bi-check-lg"></i> Save</button></div>
</div></div></div>

<!-- Add Customer Modal -->
<div class="modal fade" id="addCustomerModal" tabindex="-1"><div class="modal-dialog"><div class="modal-content">
<div class="modal-header bg-success text-white"><h5><i class="bi bi-person-plus"></i> Add New Customer</h5><button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button></div>
<div class="modal-body">
<div class="mb-3"><label class="form-label">Customer Name *</label><input type="text" class="form-control" id="newCustName" placeholder="e.g., HUD - Section 8"></div>
<div class="mb-3"><label class="form-label">Default GL Code</label>
<select class="form-select" id="newCustGL"><option value="">-- Select --</option>{% for g in gl_codes %}<option value="{{ g.code }}">{{ g.code }} - {{ g.name[:30] }}</option>{% endfor %}</select></div>
<div class="mb-3"><label class="form-label">Default Fund/Class</label>
<select class="form-select" id="newCustFund">{% for f in fund_codes %}<option value="{{ f.code }}">{{ f.name }}</option>{% endfor %}</select></div>
</div>
<div class="modal-footer"><button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Cancel</button><button type="button" class="btn btn-success" onclick="addCustomer()"><i class="bi bi-plus"></i> Add</button></div>
</div></div></div>

<!-- Add Vendor Modal -->
<div class="modal fade" id="addVendorModal" tabindex="-1"><div class="modal-dialog"><div class="modal-content">
<div class="modal-header bg-info text-white"><h5><i class="bi bi-building-add"></i> Add New Vendor</h5><button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button></div>
<div class="modal-body">
<div class="mb-3"><label class="form-label">Vendor Name *</label><input type="text" class="form-control" id="newVendorName" placeholder="e.g., Office Depot"></div>
<div class="mb-3"><label class="form-label">Default GL Code</label>
<select class="form-select" id="newVendorGL"><option value="">-- Select --</option>{% for g in gl_codes %}<option value="{{ g.code }}">{{ g.code }} - {{ g.name[:30] }}</option>{% endfor %}</select></div>
<div class="mb-3"><label class="form-label">Default Fund/Class</label>
<select class="form-select" id="newVendorFund">{% for f in fund_codes %}<option value="{{ f.code }}">{{ f.name }}</option>{% endfor %}</select></div>
</div>
<div class="modal-footer"><button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Cancel</button><button type="button" class="btn btn-info" onclick="addVendor()"><i class="bi bi-plus"></i> Add</button></div>
</div></div></div>

<!-- Add Transaction Modal -->
<div class="modal fade" id="addTransactionModal" tabindex="-1"><div class="modal-dialog modal-lg"><div class="modal-content">
<div class="modal-header bg-primary text-white"><h5><i class="bi bi-plus-circle"></i> Add New Transaction</h5><button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button></div>
<div class="modal-body">
<div class="row">
<div class="col-md-6 mb-3"><label class="form-label">Date *</label><input type="date" class="form-control" id="newTxnDate" required></div>
<div class="col-md-6 mb-3"><label class="form-label">Amount *</label>
<div class="input-group"><span class="input-group-text">$</span><input type="number" class="form-control" id="newTxnAmount" step="0.01" placeholder="0.00" required>
<select class="form-select" id="newTxnType" style="max-width:120px"><option value="deposit">Deposit (+)</option><option value="withdrawal">Withdrawal (-)</option></select></div></div>
</div>
<div class="mb-3"><label class="form-label">Description *</label><input type="text" class="form-control" id="newTxnDesc" placeholder="e.g., ACH Corporate Deposit" required></div>
<div class="row">
<div class="col-md-4 mb-3"><label class="form-label">Module</label>
<select class="form-select" id="newTxnModule"><option value="CR">Cash Receipt (CR)</option><option value="CD">Cash Disbursement (CD)</option><option value="JV">Journal Voucher (JV)</option></select></div>
<div class="col-md-4 mb-3"><label class="form-label">GL Code</label>
<select class="form-select" id="newTxnGL"><option value="">-- Select --</option>{% for g in gl_codes %}<option value="{{ g.code }}">{{ g.code }} - {{ g.name[:25] }}</option>{% endfor %}</select></div>
<div class="col-md-4 mb-3"><label class="form-label">Fund/Class</label>
<select class="form-select" id="newTxnFund"><option value="">-- Select --</option>{% for f in fund_codes %}<option value="{{ f.code }}">{{ f.name }}</option>{% endfor %}</select></div>
</div>
</div>
<div class="modal-footer"><button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Cancel</button><button type="button" class="btn btn-primary" onclick="addTransaction()"><i class="bi bi-plus"></i> Add Transaction</button></div>
</div></div></div>

</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
<script>
const glCodes={{ gl_codes|tojson }}, fundCodes={{ fund_codes|tojson }}, vendors={{ vendors|tojson }}, customers={{ customers|tojson }};
const editModal=new bootstrap.Modal(document.getElementById('editModal'));
let currentSugg={}, selectedIdx=[];

function onPayeeSelect() {
    const sel = document.getElementById('editPayeeSelect');
    const opt = sel.options[sel.selectedIndex];
    if (opt && opt.dataset.gl) document.getElementById('editGL').value = opt.dataset.gl;
    if (opt && opt.dataset.fund) document.getElementById('editFund').value = opt.dataset.fund;
    if (opt && opt.value) document.getElementById('editPayee').value = opt.value;
}

document.querySelectorAll('.edit-btn').forEach(btn=>{
  btn.onclick=function(){
    const idx=this.dataset.idx, desc=this.dataset.desc||'', mod=this.dataset.module||'CR', gl=this.dataset.gl||'', fund=this.dataset.fund||'', payee=this.dataset.payee||'', amt=parseFloat(this.dataset.amt)||0;
    document.getElementById('editIdx').value=idx;
    document.getElementById('editDesc').textContent=desc;
    document.getElementById('editAmt').textContent=(amt>=0?'+':'')+amt.toLocaleString('en-US',{style:'currency',currency:'USD'});
    document.getElementById('editModule').value=mod;
    document.getElementById('editGL').value=gl;
    document.getElementById('editFund').value=fund;
    document.getElementById('editPayee').value=payee;
    // Try to select matching payee in dropdown
    const payeeSelect=document.getElementById('editPayeeSelect');
    payeeSelect.value=payee;
    if(!payeeSelect.value && payee) payeeSelect.value='';
    updateModExplain(amt,desc);
    genSuggestion(desc,amt);
    findVendors(desc);
    editModal.show();
  };
});

function updateModExplain(amt,desc){
  const el=document.getElementById('modExplain');
  if(amt>0) el.textContent='Tip: Deposit - typically CR';
  else if(desc.toLowerCase().includes('check')) el.textContent='Tip: Check - typically CD';
  else if(desc.toLowerCase().includes('fee')) el.textContent='Tip: Fee - typically JV';
  else el.textContent='Tip: Withdrawal - typically CD';
}

function genSuggestion(desc,amt){
  const d=desc.toLowerCase();
  let mod=amt>0?'CR':'CD', gl='5510', fund='General', reason='Default';
  if(d.includes('nahasda')||d.includes('hud')){gl='3001';fund='NAHASDA';mod='CR';reason='Federal Grant';}
  else if(d.includes('bgca')||d.includes('boys and girls')){gl='3001';fund='BGC';mod='CR';reason='BGC Grant';}
  else if(d.includes('bcbs')){gl='4080';fund='BCBS';mod='CR';reason='BCBS Contribution';}
  else if(d.includes('delta dental')){gl='24060';mod='CD';reason='Health Insurance';}
  else if(d.includes('gfl')||d.includes('waste')){gl='5950';fund='Indirect';mod='CD';reason='Waste Disposal';}
  else if(d.includes('quality office')||d.includes('copier')){gl='5380';mod='CD';reason='Copier Maintenance';}
  else if(d.includes('proshred')){gl='5810';fund='Indirect';mod='CD';reason='Supplies';}
  else if(d.includes('payroll')||d.includes('adp')){gl='6600';mod='CD';reason='Payroll';}
  else if(d.includes('irs')||d.includes('eftps')||d.includes('tax')){gl='6700';mod='CD';reason='Payroll Taxes';}
  else if(d.includes('fee')||d.includes('service charge')||d.includes('bank')){gl='5060';mod='JV';reason='Bank Charges';}
  else if(d.includes('interest')&&amt>0){gl='9010';mod='CR';reason='Interest Income';}
  else if(amt>0){gl='4080';mod='CR';reason='Contribution/Deposit';}
  currentSugg={module:mod,gl_code:gl,fund_code:fund};
  const glName=glCodes.find(g=>g.code===gl)?.name||'';
  document.getElementById('suggText').innerHTML=`Module: <b>${mod}</b> | GL: <b>${gl}</b> (${glName}) | Fund: <b>${fund}</b><br><small class="text-muted">${reason}</small>`;
}

function applySugg(){
  if(currentSugg.module) document.getElementById('editModule').value=currentSugg.module;
  if(currentSugg.gl_code) document.getElementById('editGL').value=currentSugg.gl_code;
  if(currentSugg.fund_code) document.getElementById('editFund').value=currentSugg.fund_code;
}

function findVendors(desc){
  const d=desc.toLowerCase(), panel=document.getElementById('vendorPanel'), div=document.getElementById('vendorMatches');
  const allEntities = [...vendors, ...customers];
  const matches=allEntities.filter(v=>v.name.toLowerCase().split(' ').some(w=>w.length>2&&d.includes(w)));
  if(matches.length){
    panel.classList.remove('d-none');
    div.innerHTML=matches.slice(0,3).map(v=>`<div class="vendor-item" onclick="applyVendor('${v.gl_code}','${v.fund}','${v.name}')"><strong>${v.name}</strong> <span class="badge bg-secondary">${v.type||'Vendor'}</span><br><small>GL:${v.gl_code} Fund:${v.fund}</small></div>`).join('');
  } else panel.classList.add('d-none');
}

function applyVendor(gl,fund,name){
  document.getElementById('editGL').value=gl;
  document.getElementById('editFund').value=fund;
  document.getElementById('editPayee').value=name;
}

document.getElementById('saveEdit').onclick=function(){
  fetch('/update_transaction',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({
      index:parseInt(document.getElementById('editIdx').value),
      module:document.getElementById('editModule').value,
      gl_code:document.getElementById('editGL').value,
      fund_code:document.getElementById('editFund').value,
      payee:document.getElementById('editPayee').value || document.getElementById('editPayeeSelect').value
    })
  }).then(r=>r.json()).then(d=>{if(d.status==='success')location.reload();else alert(d.message);});
};

function addCustomer() {
  const name = document.getElementById('newCustName').value.trim();
  const gl = document.getElementById('newCustGL').value;
  const fund = document.getElementById('newCustFund').value;
  if (!name) { alert('Customer name is required'); return; }
  fetch('/add_customer', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({name, gl_code: gl, fund})
  }).then(r=>r.json()).then(d=>{alert(d.message);if(d.status==='success')location.reload();});
}

function addVendor() {
  const name = document.getElementById('newVendorName').value.trim();
  const gl = document.getElementById('newVendorGL').value;
  const fund = document.getElementById('newVendorFund').value;
  if (!name) { alert('Vendor name is required'); return; }
  fetch('/add_vendor', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({name, gl_code: gl, fund})
  }).then(r=>r.json()).then(d=>{alert(d.message);if(d.status==='success')location.reload();});
}

function addTransaction() {
  const date = document.getElementById('newTxnDate').value;
  const amount = parseFloat(document.getElementById('newTxnAmount').value);
  const type = document.getElementById('newTxnType').value;
  const desc = document.getElementById('newTxnDesc').value.trim();
  const module = document.getElementById('newTxnModule').value;
  const gl = document.getElementById('newTxnGL').value;
  const fund = document.getElementById('newTxnFund').value;
  if (!date || !amount || !desc) { alert('Date, Amount, and Description are required'); return; }
  const finalAmount = type === 'withdrawal' ? -Math.abs(amount) : Math.abs(amount);
  fetch('/add_transaction', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({date, amount: finalAmount, description: desc, module, gl_code: gl, fund_code: fund})
  }).then(r=>r.json()).then(d=>{alert(d.message);if(d.status==='success')location.reload();});
}

document.querySelectorAll('.delete-btn').forEach(btn=>{
  btn.onclick=function(){
    const idx = this.dataset.idx;
    const desc = this.dataset.desc || 'this transaction';
    if(!confirm('Are you sure you want to delete "' + desc.substring(0, 50) + '"?')) return;
    fetch('/delete_transaction', {method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({index: parseInt(idx)})
    }).then(r=>r.json()).then(d=>{if(d.status==='success')location.reload();else alert(d.message);});
  };
});

function toggleAll(){
  const all=document.getElementById('selAll').checked;
  document.querySelectorAll('.rowCb').forEach(cb=>{cb.checked=all;cb.closest('tr').classList.toggle('selected',all);});
  updateSel();
}
function updateSel(){
  selectedIdx=Array.from(document.querySelectorAll('.rowCb:checked')).map(cb=>parseInt(cb.value));
  document.getElementById('selCount').textContent=selectedIdx.length;
  document.getElementById('bulkBar').classList.toggle('d-none',selectedIdx.length===0);
  document.querySelectorAll('.rowCb').forEach(cb=>cb.closest('tr').classList.toggle('selected',cb.checked));
}
function clearSel(){
  document.querySelectorAll('.rowCb').forEach(cb=>{cb.checked=false;cb.closest('tr').classList.remove('selected');});
  document.getElementById('selAll').checked=false;
  selectedIdx=[];
  document.getElementById('bulkBar').classList.add('d-none');
}
function applyBulk(){
  if(!selectedIdx.length){alert('Select transactions first');return;}
  const updates={};
  const m=document.getElementById('bulkModule').value,g=document.getElementById('bulkGL').value,f=document.getElementById('bulkFund').value;
  if(m)updates.module=m;if(g)updates.gl_code=g;if(f)updates.fund_code=f;
  if(!Object.keys(updates).length){alert('Select a field to update');return;}
  fetch('/bulk_update',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({indices:selectedIdx,updates})})
  .then(r=>r.json()).then(d=>{if(d.status==='success')location.reload();else alert(d.message);});
}
function undoLast(){
  fetch('/undo_last',{method:'POST'}).then(r=>r.json()).then(d=>{alert(d.message);if(d.status==='success')location.reload();});
}

// Filter by module when KPI card is clicked
function filterByModule(module){
  // Scroll to transaction section
  const tabs = document.querySelector('#transactionTabs');
  if(tabs) {
    tabs.scrollIntoView({behavior: 'smooth'});
  }
  // Click the corresponding tab
  const tabId = module.toLowerCase();
  const tabBtn = document.querySelector('[data-bs-target="#' + tabId + '"]');
  if(tabBtn){
    tabBtn.click();
  }
}

// Initialize Select2 for searchable dropdowns
$(document).ready(function() {
    // Edit Modal dropdowns
    $('#editGL').select2({
        placeholder: 'Search GL Code...',
        allowClear: true,
        dropdownParent: $('#editModal'),
        width: '100%'
    });
    
    $('#editFund').select2({
        placeholder: 'Search Fund...',
        allowClear: true,
        dropdownParent: $('#editModal'),
        width: '100%'
    });
    
    $('#editPayeeSelect').select2({
        placeholder: 'Search Customer/Vendor...',
        allowClear: true,
        dropdownParent: $('#editModal'),
        width: '100%'
    });
    
    // Bulk action dropdowns
    $('#bulkGL').select2({
        placeholder: 'GL Code...',
        allowClear: true,
        width: '200px'
    });
    
    $('#bulkFund').select2({
        placeholder: 'Fund...',
        allowClear: true,
        width: '200px'
    });
    
    // Add Customer Modal dropdowns
    $('#newCustGL').select2({
        placeholder: 'Search GL Code...',
        allowClear: true,
        dropdownParent: $('#addCustomerModal'),
        width: '100%'
    });
    
    $('#newCustFund').select2({
        placeholder: 'Search Fund...',
        allowClear: true,
        dropdownParent: $('#addCustomerModal'),
        width: '100%'
    });
    
    // Add Vendor Modal dropdowns
    $('#newVendorGL').select2({
        placeholder: 'Search GL Code...',
        allowClear: true,
        dropdownParent: $('#addVendorModal'),
        width: '100%'
    });
    
    $('#newVendorFund').select2({
        placeholder: 'Search Fund...',
        allowClear: true,
        dropdownParent: $('#addVendorModal'),
        width: '100%'
    });
    
    // Re-initialize Select2 when modal opens (fixes display issues)
    $('#editModal').on('shown.bs.modal', function() {
        $('#editGL').select2({
            placeholder: 'Search GL Code...',
            allowClear: true,
            dropdownParent: $('#editModal'),
            width: '100%'
        });
        $('#editFund').select2({
            placeholder: 'Search Fund...',
            allowClear: true,
            dropdownParent: $('#editModal'),
            width: '100%'
        });
        $('#editPayeeSelect').select2({
            placeholder: 'Search Customer/Vendor...',
            allowClear: true,
            dropdownParent: $('#editModal'),
            width: '100%'
        });
    });
    
    // Re-initialize Select2 when Add Customer modal opens
    $('#addCustomerModal').on('shown.bs.modal', function() {
        $('#newCustGL').select2({
            placeholder: 'Search GL Code...',
            allowClear: true,
            dropdownParent: $('#addCustomerModal'),
            width: '100%'
        });
        $('#newCustFund').select2({
            placeholder: 'Search Fund...',
            allowClear: true,
            dropdownParent: $('#addCustomerModal'),
            width: '100%'
        });
    });
    
    // Re-initialize Select2 when Add Vendor modal opens
    $('#addVendorModal').on('shown.bs.modal', function() {
        $('#newVendorGL').select2({
            placeholder: 'Search GL Code...',
            allowClear: true,
            dropdownParent: $('#addVendorModal'),
            width: '100%'
        });
        $('#newVendorFund').select2({
            placeholder: 'Search Fund...',
            allowClear: true,
            dropdownParent: $('#addVendorModal'),
            width: '100%'
        });
    });
    
    // Handle Customer/Vendor selection to auto-fill GL and Fund
    $('#editPayeeSelect').on('select2:select', function(e) {
        var data = e.params.data;
        var element = $(data.element);
        var gl = element.data('gl');
        var fund = element.data('fund');
        if (gl) $('#editGL').val(gl).trigger('change');
        if (fund) $('#editFund').val(fund).trigger('change');
    });
});
</script></body></html>'''

RESULTS_TEMPLATE = '''<!DOCTYPE html>
<html><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Results - Bank Transaction Posting Tool</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.0/font/bootstrap-icons.css" rel="stylesheet">
<style>.audit-item{font-size:.85em;padding:5px 10px;border-left:3px solid #6c757d;margin-bottom:5px;background:#f8f9fa}</style>
</head><body>
<nav class="navbar navbar-dark bg-primary"><div class="container"><a class="navbar-brand" href="/"><i class="bi bi-bank"></i> Bank Transaction Posting Tool</a></div></nav>
<div class="container mt-4">
<div class="card bg-success text-white mb-4"><div class="card-body text-center"><h3><i class="bi bi-check-circle"></i> Processing Complete!</h3></div></div>
<div class="row"><div class="col-md-8">
<div class="card"><div class="card-header d-flex justify-content-between align-items-center"><h5 class="mb-0"><i class="bi bi-download"></i> Download Files</h5>
{% if output_files %}<a href="/download_all_zip" class="btn btn-success btn-sm"><i class="bi bi-file-earmark-zip"></i> Download All as ZIP</a>{% endif %}</div>
<div class="card-body">{% if output_files %}<div class="list-group">{% for f in output_files %}
<a href="/download/{{ f.filename }}" class="list-group-item list-group-item-action d-flex justify-content-between"><div><i class="bi bi-file-earmark-excel text-success"></i> <strong>{{ f.filename }}</strong><br><small class="text-muted">{{ f.description }}</small></div><span class="badge bg-primary">Download</span></a>
{% endfor %}</div>{% else %}<p class="text-muted">No files</p>{% endif %}</div></div></div>
<div class="col-md-4">
<div class="card"><div class="card-header"><h5><i class="bi bi-clock-history"></i> Audit Trail</h5></div>
<div class="card-body" style="max-height:250px;overflow-y:auto">{% if audit_trail %}{% for e in audit_trail|reverse %}<div class="audit-item"><small>{{ e.timestamp }}</small> - <strong>{{ e.transaction }}</strong><br><small>{{ e.changes|join(', ') }}</small></div>{% endfor %}{% else %}<p class="text-muted">No edits</p>{% endif %}</div></div>
<div class="card mt-3"><div class="card-body text-center"><a href="/" class="btn btn-primary"><i class="bi bi-upload"></i> Process Another</a></div></div>
</div></div></div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body></html>'''

if __name__ == '__main__':
    from config import FLASK_HOST, FLASK_PORT, FLASK_DEBUG

    print("="*70)
    print("  Bank Transaction Posting Tool - MongoDB Enhanced")
    print("="*70)
    print(f"  Web Interface: http://{FLASK_HOST}:{FLASK_PORT}")
    print(f"  API Endpoint:  http://{FLASK_HOST}:{FLASK_PORT}/api/")
    print("-"*70)
    print(f"  MongoDB URI: {MONGODB_URI}")
    print(f"  Database: {MONGODB_DATABASE}")
    print(f"  Connection: Will initialize on first use (non-blocking)")
    print("-"*70)
    print("  API Endpoints:")
    print("    GET  /api/status          - Health check")
    print("    GET  /api/stats           - Dashboard statistics")
    print("    ---  TRANSACTIONS  ---")
    print("    GET  /api/transactions    - List transactions")
    print("    POST /api/transactions    - Create transaction")
    print("    GET  /api/transactions/<id> - Get transaction")
    print("    PUT  /api/transactions/<id> - Update transaction")
    print("    DELETE /api/transactions/<id> - Delete transaction")
    print("    ---  BATCHES  ---")
    print("    GET  /api/batches         - List batches")
    print("    POST /api/batches         - Create batch with transactions")
    print("    GET  /api/batches/<id>    - Get batch with transactions")
    print("    POST /api/batches/<id>/process - Classify batch")
    print("    ---  MASTER DATA  ---")
    print("    GET  /api/gl-codes        - List GL codes")
    print("    POST /api/gl-codes        - Create GL code")
    print("    GET  /api/fund-codes      - List fund codes")
    print("    GET  /api/vendors         - List vendors")
    print("    POST /api/vendors         - Create vendor")
    print("    GET  /api/customers       - List customers")
    print("    POST /api/customers       - Create customer")
    print("    ---  UTILITIES  ---")
    print("    GET  /api/audit-logs      - View audit logs")
    print("    POST /api/sync/master-data - Sync to MongoDB")
    print("="*70)
    print("\n[*] Starting server...")
    app.run(debug=FLASK_DEBUG, host=FLASK_HOST, port=FLASK_PORT)